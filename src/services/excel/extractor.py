# SPDX-License-Identifier: AGPL-3.0-or-later
# Copyright (C) 2026 Assistance Micro Design
"""Extracteur de données pour fichiers Excel."""

from __future__ import annotations

import logging
import re
from datetime import datetime
from pathlib import Path
from typing import TYPE_CHECKING, Any

import xlrd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from src.models.excel import (
    CellType,
    ExcelCell,
    ExcelDocument,
    ExcelFormula,
    ExcelSheet,
    ExcelTable,
)
from src.models.types import CellValue


if TYPE_CHECKING:
    from openpyxl.cell import Cell


logger = logging.getLogger(__name__)


class ExcelExtractor:
    """Extracteur de donnees pour fichiers Excel.

    Attributes:
        data_only: Si True, charge les resultats des formules au lieu des formules brutes.
    """

    def __init__(self, data_only: bool = False) -> None:
        """Initialise l'extracteur.

        Args:
            data_only: Si True, charge les résultats des formules.
                      Si False, charge les formules brutes.
        """
        self.data_only = data_only

    async def extract(self, file_path: str | Path) -> ExcelDocument:
        """Extrait les données d'un fichier Excel.

        Args:
            file_path: Chemin vers le fichier Excel.

        Returns:
            ExcelDocument avec toutes les données extraites.

        Raises:
            FileNotFoundError: Si le fichier n'existe pas.
            ValueError: Si le format n'est pas supporté.
        """
        path = Path(file_path)

        if not path.exists():
            msg = f"Fichier non trouvé: {path}"
            raise FileNotFoundError(msg)

        ext = path.suffix.lower()

        if ext == ".xlsx":
            return await self._extract_xlsx(path)
        if ext == ".xls":
            return await self._extract_xls(path)

        msg = f"Format non supporté: {ext}"
        raise ValueError(msg)

    async def _extract_xlsx(self, path: Path) -> ExcelDocument:
        """Extrait les données d'un fichier .xlsx avec openpyxl."""
        # Charger deux fois : une pour les formules, une pour les valeurs
        wb_formulas = load_workbook(path, data_only=False)
        wb_values = load_workbook(path, data_only=True)

        sheets: list[ExcelSheet] = []
        total_formulas = 0
        total_tables = 0

        for idx, sheet_name in enumerate(wb_formulas.sheetnames):
            ws_formulas = wb_formulas[sheet_name]
            ws_values = wb_values[sheet_name]

            # Type guard: ensure we have Worksheet, not ReadOnlyWorksheet
            if not isinstance(ws_formulas, Worksheet) or not isinstance(ws_values, Worksheet):
                logger.warning("Feuille %s n'est pas un Worksheet standard, ignorée", sheet_name)
                continue

            sheet = await self._extract_sheet(ws_formulas, ws_values, sheet_name, idx)
            sheets.append(sheet)
            total_formulas += len(sheet.formulas)
            total_tables += len(sheet.tables)

        # Propriétés du document
        props: dict[str, Any] = {}
        if wb_formulas.properties:
            props = {
                "title": wb_formulas.properties.title,
                "author": wb_formulas.properties.creator,
                "created": (
                    str(wb_formulas.properties.created) if wb_formulas.properties.created else None
                ),
                "modified": (
                    str(wb_formulas.properties.modified)
                    if wb_formulas.properties.modified
                    else None
                ),
            }

        # Plages nommées
        named_ranges: dict[str, str] = {}
        if wb_formulas.defined_names:
            for defined_name in wb_formulas.defined_names.definedName:
                named_ranges[defined_name.name] = str(defined_name.attr_text)

        wb_formulas.close()
        wb_values.close()

        return ExcelDocument(
            filename=path.name,
            file_path=str(path.absolute()),
            format="xlsx",
            sheets=sheets,
            named_ranges=named_ranges,
            properties=props,
            total_formulas=total_formulas,
            total_tables=total_tables,
        )

    async def _extract_sheet(
        self,
        ws_formulas: Worksheet,
        ws_values: Worksheet,
        name: str,
        index: int,
    ) -> ExcelSheet:
        """Extrait une feuille Excel."""
        cells: list[list[ExcelCell]] = []
        formulas: list[ExcelFormula] = []
        tables: list[ExcelTable] = []

        # Dimensions
        max_row = ws_formulas.max_row or 0
        max_col = ws_formulas.max_column or 0

        # Extraire cellules
        for row_idx in range(1, max_row + 1):
            row_cells: list[ExcelCell] = []
            for col_idx in range(1, max_col + 1):
                cell_formula = ws_formulas.cell(row=row_idx, column=col_idx)
                cell_value = ws_values.cell(row=row_idx, column=col_idx)

                excel_cell = self._process_cell(cell_formula, cell_value, row_idx, col_idx)
                row_cells.append(excel_cell)

                # Collecter les formules
                if excel_cell.formula:
                    formulas.append(
                        ExcelFormula(
                            cell=f"{excel_cell.column_letter}{row_idx}",
                            sheet=name,
                            formula=excel_cell.formula,
                            result=excel_cell.value,
                            dependencies=self._parse_formula_dependencies(excel_cell.formula),
                        )
                    )

            cells.append(row_cells)

        # Détecter les tableaux
        tables = self._detect_tables(ws_formulas, ws_values)

        # Cellules fusionnées
        merged = [str(m) for m in ws_formulas.merged_cells.ranges]

        return ExcelSheet(
            name=name,
            index=index,
            rows_count=max_row,
            columns_count=max_col,
            cells=cells,
            tables=tables,
            formulas=formulas,
            merged_cells=merged,
        )

    def _detect_cell_type(
        self,
        value: CellValue,
        formula_value: str | None,
    ) -> tuple[CellType, CellValue, str | None]:
        """Detecte le type de cellule et normalise la valeur.

        Args:
            value: Valeur de la cellule.
            formula_value: Valeur brute (peut contenir formule).

        Returns:
            Tuple (cell_type, normalized_value, formula).
        """
        # Formule Excel
        if isinstance(formula_value, str) and formula_value.startswith("="):
            return CellType.FORMULA, value, formula_value

        # Cellule vide
        if value is None:
            return CellType.EMPTY, None, None

        # Boolean (avant number car bool est subclass de int)
        if isinstance(value, bool):
            return CellType.BOOLEAN, value, None

        # Nombre
        if isinstance(value, (int, float)):
            return CellType.NUMBER, value, None

        # Date
        if isinstance(value, datetime):
            return CellType.DATE, value.isoformat(), None

        return CellType.TEXT, str(value) if value else None, None

    def _process_cell(
        self,
        cell_formula: Cell,
        cell_value: Cell,
        row: int,
        col: int,
    ) -> ExcelCell:
        """Traite une cellule et détermine son type."""
        col_letter = get_column_letter(col)
        value: CellValue = cell_value.value

        cell_type, value, formula = self._detect_cell_type(value, cell_formula.value)

        return ExcelCell(
            row=row,
            column=col,
            column_letter=col_letter,
            value=value,
            formula=formula,
            cell_type=cell_type,
        )

    def _detect_tables(
        self,
        ws_formulas: Worksheet,
        ws_values: Worksheet,
    ) -> list[ExcelTable]:
        """Détecte les tableaux dans une feuille."""
        tables: list[ExcelTable] = []

        # Tableaux Excel officiels
        for table in ws_formulas.tables.values():
            headers: list[str] = []
            data: list[list[CellValue]] = []

            # Extraire les données du tableau
            table_range = ws_values[table.ref]

            # Handle single cell reference
            if not hasattr(table_range, "__iter__"):
                data.append([table_range.value])

            # Handle range: iterate rows
            rows = table_range if hasattr(table_range, "__iter__") else ()
            for i, row in enumerate(rows):
                row_values: list[CellValue] = (
                    [cell.value for cell in row] if hasattr(row, "__iter__") else [row.value]
                )

                if i == 0 and table.headerRowCount:
                    headers = [str(v) if v else "" for v in row_values]
                    continue
                data.append(row_values)

            tables.append(
                ExcelTable(
                    name=table.name,
                    range=table.ref,
                    headers=headers,
                    data=data,
                )
            )

        return tables

    def _parse_formula_dependencies(self, formula: str) -> list[str]:
        """Parse les références de cellules dans une formule."""
        # Pattern pour les références de cellules (A1, $A$1, A1:B10, etc.)
        pattern = r"\$?[A-Z]+\$?\d+"
        matches = re.findall(pattern, formula.upper())

        return list(set(matches))

    def _detect_xls_cell_type(
        self,
        value: CellValue,
        cell_type_xlrd: int,
        datemode: int,
    ) -> tuple[CellType, CellValue]:
        """Detecte le type de cellule pour fichier .xls.

        Args:
            value: Valeur de la cellule.
            cell_type_xlrd: Type xlrd de la cellule.
            datemode: Mode de date du workbook.

        Returns:
            Tuple (cell_type, normalized_value).
        """
        if cell_type_xlrd == xlrd.XL_CELL_EMPTY:
            return CellType.EMPTY, value

        if cell_type_xlrd == xlrd.XL_CELL_NUMBER:
            return CellType.NUMBER, value

        if cell_type_xlrd == xlrd.XL_CELL_DATE:
            # Convertir la date Excel
            if isinstance(value, float):
                value = xlrd.xldate_as_datetime(value, datemode).isoformat()
            return CellType.DATE, value

        if cell_type_xlrd == xlrd.XL_CELL_BOOLEAN:
            return CellType.BOOLEAN, bool(value)

        if cell_type_xlrd == xlrd.XL_CELL_ERROR:
            return CellType.ERROR, value

        return CellType.TEXT, str(value) if value else None

    async def _extract_xls(self, path: Path) -> ExcelDocument:
        """Extrait les données d'un fichier .xls legacy avec xlrd."""
        wb = xlrd.open_workbook(str(path))
        sheets: list[ExcelSheet] = []

        for idx in range(wb.nsheets):
            ws = wb.sheet_by_index(idx)

            cells: list[list[ExcelCell]] = []
            tables: list[ExcelTable] = []

            for row_idx in range(ws.nrows):
                row_cells: list[ExcelCell] = []
                for col_idx in range(ws.ncols):
                    value: CellValue = ws.cell_value(row_idx, col_idx)
                    cell_type_xlrd = ws.cell_type(row_idx, col_idx)

                    # Convertir type xlrd vers notre enum
                    cell_type, value = self._detect_xls_cell_type(
                        value, cell_type_xlrd, wb.datemode
                    )

                    row_cells.append(
                        ExcelCell(
                            row=row_idx + 1,
                            column=col_idx + 1,
                            column_letter=get_column_letter(col_idx + 1),
                            value=value,
                            formula=None,  # xlrd ne supporte pas les formules
                            cell_type=cell_type,
                        )
                    )

                cells.append(row_cells)

            sheets.append(
                ExcelSheet(
                    name=ws.name,
                    index=idx,
                    rows_count=ws.nrows,
                    columns_count=ws.ncols,
                    cells=cells,
                    tables=tables,
                    formulas=[],  # Pas de formules avec xlrd
                    merged_cells=[],
                )
            )

        return ExcelDocument(
            filename=path.name,
            file_path=str(path.absolute()),
            format="xls",
            sheets=sheets,
            named_ranges={},
            properties={},
            total_formulas=0,
            total_tables=0,
        )
