# SPDX-License-Identifier: AGPL-3.0-or-later
# Copyright (C) 2025-2026 Assistance Micro Design
"""Editeur de fichiers Excel (.xlsx) existants."""

from __future__ import annotations

import asyncio
import logging
from pathlib import Path
from typing import TYPE_CHECKING, Any

from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from src.core.config import settings
from src.core.exceptions import (
    ExcelChartError,
    ExcelFileNotFoundError,
    ExcelSheetNotFoundError,
)
from src.core.file_validation import validate_decompressed_size
from src.models.api import EditExcelParams, EditExcelResult
from src.models.excel_edit import (
    AddChartOp,
    AddDataValidationOp,
    AddSheetOp,
    ApplyStylesOp,
    DeleteRowsOp,
    DeleteSheetOp,
    InsertRowsOp,
    MergeCellsOp,
    RemoveChartsOp,
    RenameSheetOp,
    SetSheetPropertiesOp,
    UnmergeCellsOp,
    UpdateCellsOp,
)
from src.services.excel.generator import ExcelGenerator


if TYPE_CHECKING:
    from collections.abc import Callable

    from openpyxl import Workbook
    from openpyxl.worksheet.worksheet import Worksheet

logger = logging.getLogger(__name__)


class ExcelEditor:
    """Editeur de fichiers Excel existants.

    Utilise ExcelGenerator en composition pour reutiliser
    les builders de styles et les methodes de charts.

    Attributes:
        _output_path: Repertoire contenant les fichiers a editer.
        _generator: Instance ExcelGenerator pour reutiliser les helpers.
    """

    def __init__(self, output_path: Path | None = None) -> None:
        """Initialise l'editeur.

        Args:
            output_path: Repertoire des fichiers (defaut: settings.OUTPUT_PATH).
        """
        self._output_path = Path(output_path or settings.OUTPUT_PATH)
        self._generator = ExcelGenerator(output_path=self._output_path)

    def ensure_output_dir(self) -> None:
        """Delegue la creation du repertoire de sortie au generator."""
        self._generator.ensure_output_dir()

    async def edit(self, params: EditExcelParams) -> EditExcelResult:
        """Point d'entree principal. Edite le fichier xlsx.

        Args:
            params: Parametres d'edition (filename + operations).

        Returns:
            Resultat avec stats des operations.

        Raises:
            ExcelFileNotFoundError: Si le fichier n'existe pas dans OUTPUT_PATH.
            ExcelSheetNotFoundError: Si une feuille referencee n'existe pas.
            ExcelChartError: Loguee en warning, l'operation est comptee comme skipped.
        """
        return await asyncio.to_thread(self._edit_sync, params)

    def _edit_sync(self, params: EditExcelParams) -> EditExcelResult:
        """Edition synchrone du fichier (appelee via to_thread).

        Charge le workbook, applique les operations et sauvegarde.
        """
        safe_filename = self._generator.sanitize_filename(params.filename)
        file_path = self._output_path / safe_filename
        if not file_path.exists():
            raise ExcelFileNotFoundError(safe_filename)

        validate_decompressed_size(file_path, settings.MAX_DECOMPRESSED_MB)
        wb = load_workbook(str(file_path))
        applied, skipped = self._apply_operations(wb, params.operations)
        file_size = self._generator.save_and_verify(wb, file_path, safe_filename)

        logger.info(
            "Excel edite: %s (%d ops, %d ignorees, %d octets)",
            safe_filename,
            applied,
            skipped,
            file_size,
        )

        return EditExcelResult(
            file_path=str(file_path),
            filename=safe_filename,
            operations_applied=applied,
            operations_skipped=skipped,
            file_size_bytes=file_size,
        )

    def _apply_operations(self, wb: Workbook, operations: list[Any]) -> tuple[int, int]:
        """Applique les operations en sequence via dispatch dict.

        Les ExcelChartError sont catchees et comptees comme skipped.

        Args:
            wb: Workbook openpyxl ouvert.
            operations: Liste d'operations typees (discriminated union).

        Returns:
            Tuple (applied, skipped).
        """
        op_handlers: dict[str, Callable[..., None]] = {
            "update_cells": self._op_update_cells,
            "insert_rows": self._op_insert_rows,
            "delete_rows": self._op_delete_rows,
            "apply_styles": self._op_apply_styles,
            "add_sheet": self._op_add_sheet,
            "delete_sheet": self._op_delete_sheet,
            "rename_sheet": self._op_rename_sheet,
            "add_chart": self._op_add_chart,
            "remove_charts": self._op_remove_charts,
            "add_data_validation": self._op_add_data_validation,
            "merge_cells": self._op_merge_cells,
            "unmerge_cells": self._op_unmerge_cells,
            "set_sheet_properties": self._op_set_sheet_properties,
        }
        applied = 0
        skipped = 0
        for op in operations:
            handler = op_handlers[op.op]
            try:
                handler(wb, op)
                applied += 1
            except ExcelChartError as exc:
                logger.warning("Operation chart ignoree: %s", exc)
                skipped += 1
        return applied, skipped

    # --- Helpers ---

    def _get_sheet(self, wb: Workbook, name: str) -> Worksheet:
        """Recupere une feuille par nom ou leve ExcelSheetNotFoundError."""
        if name not in wb.sheetnames:
            raise ExcelSheetNotFoundError(name, wb.sheetnames)
        return wb[name]

    # --- Operations ---

    def _op_update_cells(self, wb: Workbook, op: UpdateCellsOp) -> None:
        """Modifier des valeurs de cellules."""
        ws = self._get_sheet(wb, op.sheet)
        for cell_ref, value in op.cells.items():
            self._generator.check_cell_value_safety(value)
            ws[cell_ref] = value

    def _op_insert_rows(self, wb: Workbook, op: InsertRowsOp) -> None:
        """Inserer des lignes (append ou insert)."""
        ws = self._get_sheet(wb, op.sheet)
        for row_data in op.rows:
            for value in row_data:
                self._generator.check_cell_value_safety(value)
        if op.at_row is not None:
            ws.insert_rows(op.at_row, amount=len(op.rows))
            for row_offset, row_data in enumerate(op.rows):
                for col_idx, value in enumerate(row_data, start=1):
                    ws.cell(row=op.at_row + row_offset, column=col_idx, value=value)
            return
        for row_data in op.rows:
            ws.append(row_data)

    def _op_delete_rows(self, wb: Workbook, op: DeleteRowsOp) -> None:
        """Supprimer des lignes."""
        ws = self._get_sheet(wb, op.sheet)
        count = op.end_row - op.start_row + 1
        ws.delete_rows(op.start_row, amount=count)

    def _op_apply_styles(self, wb: Workbook, op: ApplyStylesOp) -> None:
        """Appliquer des styles via ExcelGenerator."""
        ws = self._get_sheet(wb, op.sheet)
        self._generator.apply_styles(ws, op.styles)

    def _op_add_sheet(self, wb: Workbook, op: AddSheetOp) -> None:
        """Ajouter une feuille avec donnees optionnelles."""
        if op.headers:
            for header in op.headers:
                self._generator.check_cell_value_safety(header)
        for row_data in op.rows:
            for value in row_data:
                self._generator.check_cell_value_safety(value)
        ws = wb.create_sheet(title=op.name)
        current_row = 1
        if op.headers:
            for col_idx, header in enumerate(op.headers, start=1):
                cell = ws.cell(row=current_row, column=col_idx, value=header)
                cell.font = Font(bold=True)
            current_row += 1
        for row_data in op.rows:
            for col_idx, value in enumerate(row_data, start=1):
                ws.cell(row=current_row, column=col_idx, value=value)
            current_row += 1

    def _op_delete_sheet(self, wb: Workbook, op: DeleteSheetOp) -> None:
        """Supprimer une feuille."""
        ws = self._get_sheet(wb, op.name)
        wb.remove(ws)

    def _op_rename_sheet(self, wb: Workbook, op: RenameSheetOp) -> None:
        """Renommer une feuille."""
        ws = self._get_sheet(wb, op.name)
        ws.title = op.new_name

    def _op_add_chart(self, wb: Workbook, op: AddChartOp) -> None:
        """Ajouter un graphique via ExcelGenerator."""
        ws = self._get_sheet(wb, op.sheet)
        self._generator.add_chart(ws, op.chart)

    def _op_remove_charts(self, wb: Workbook, op: RemoveChartsOp) -> None:
        """Supprimer tous les graphiques d'une feuille."""
        ws = self._get_sheet(wb, op.sheet)
        ws._charts.clear()  # openpyxl n'expose pas d'API publique pour supprimer les charts

    def _op_add_data_validation(self, wb: Workbook, op: AddDataValidationOp) -> None:
        """Ajouter une validation de donnees via ExcelGenerator."""
        ws = self._get_sheet(wb, op.sheet)
        self._generator.add_data_validation(ws, op.validation)

    def _op_merge_cells(self, wb: Workbook, op: MergeCellsOp) -> None:
        """Fusionner des cellules via ExcelGenerator."""
        ws = self._get_sheet(wb, op.sheet)
        self._generator.merge_cells(ws, op.merge)

    def _op_unmerge_cells(self, wb: Workbook, op: UnmergeCellsOp) -> None:
        """Defusionner des cellules."""
        ws = self._get_sheet(wb, op.sheet)
        ws.unmerge_cells(op.range)

    def _op_set_sheet_properties(self, wb: Workbook, op: SetSheetPropertiesOp) -> None:
        """Modifier les proprietes d'une feuille."""
        ws = self._get_sheet(wb, op.sheet)
        if op.column_widths:
            for col_letter, width in op.column_widths.items():
                ws.column_dimensions[col_letter].width = width
        if op.freeze_panes is not None:
            ws.freeze_panes = op.freeze_panes
        if op.tab_color is not None:
            ws.sheet_properties.tabColor = op.tab_color
        if op.auto_filter is not None and op.auto_filter:
            max_col = ws.max_column or 1
            max_row = ws.max_row or 1
            last_col = get_column_letter(max_col)
            ws.auto_filter.ref = f"A1:{last_col}{max_row}"
