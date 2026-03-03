# SPDX-License-Identifier: AGPL-3.0-or-later
# Copyright (C) 2026 Assistance Micro Design
"""Generateur de fichiers Excel (.xlsx)."""

from __future__ import annotations

import asyncio
import logging
from pathlib import Path
from typing import TYPE_CHECKING

from openpyxl import Workbook
from openpyxl.chart import AreaChart, BarChart, LineChart, PieChart, Reference, ScatterChart
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from src.core.config import settings
from src.core.exceptions import ExcelChartError, ExcelGenerationError, ExcelOutputTooLargeError
from src.models.api import CreateExcelParams, CreateExcelResult
from src.models.excel_generation import (
    CellStyleDef,
    ChartDef,
    DataValidationDef,
    MergedCellDef,
    SheetDef,
)


if TYPE_CHECKING:
    from openpyxl.worksheet.worksheet import Worksheet

logger = logging.getLogger(__name__)


class ExcelGenerator:
    """Genere des fichiers Excel a partir de definitions structurees.

    Utilise openpyxl pour creer des classeurs .xlsx avec donnees,
    styles, graphiques et validations de donnees.

    Attributes:
        _output_path: Repertoire de sortie des fichiers generes.
    """

    def __init__(self, output_path: Path | None = None) -> None:
        """Initialise le generateur.

        Args:
            output_path: Repertoire de sortie (defaut: settings.OUTPUT_PATH).
        """
        self._output_path = Path(output_path or settings.OUTPUT_PATH)
        self._max_output_size_mb = settings.MAX_OUTPUT_FILE_SIZE_MB

    async def generate(self, params: CreateExcelParams) -> CreateExcelResult:
        """Point d'entree principal. Cree le fichier xlsx.

        Args:
            params: Parametres de creation du fichier.

        Returns:
            Resultat avec chemin, stats et taille du fichier.
        """
        return await asyncio.to_thread(self._generate_sync, params)

    def _generate_sync(self, params: CreateExcelParams) -> CreateExcelResult:
        """Generation synchrone du fichier (appelee via to_thread)."""
        self.ensure_output_dir()
        safe_filename = self.sanitize_filename(params.filename)
        file_path = self._output_path / safe_filename
        overwritten = file_path.exists()

        wb = Workbook()
        default_sheet = wb.active
        if default_sheet is not None:
            wb.remove(default_sheet)

        total_rows, total_charts = self._process_sheets(wb, params.sheets)

        if params.author:
            wb.properties.creator = params.author

        file_size = self.save_and_verify(wb, file_path, safe_filename)

        logger.info(
            "Excel genere: %s (%d feuilles, %d lignes, %d octets)",
            safe_filename,
            len(params.sheets),
            total_rows,
            file_size,
        )

        return CreateExcelResult(
            file_path=str(file_path),
            filename=safe_filename,
            sheets_created=len(params.sheets),
            total_rows=total_rows,
            total_charts=total_charts,
            file_size_bytes=file_size,
            overwritten=overwritten,
        )

    def _process_sheets(self, wb: Workbook, sheets: list[SheetDef]) -> tuple[int, int]:
        """Cree et configure toutes les feuilles du classeur.

        Returns:
            Tuple (total_rows, total_charts).
        """
        total_rows = 0
        total_charts = 0
        for sheet_def in sheets:
            ws = self._create_sheet(wb, sheet_def)
            total_rows += len(sheet_def.rows)
            self.apply_styles(ws, sheet_def.styles)
            for chart_def in sheet_def.charts:
                try:
                    self.add_chart(ws, chart_def)
                    total_charts += 1
                except ExcelChartError as exc:
                    logger.warning(
                        "Graphique ignore '%s': %s",
                        chart_def.title or "sans titre",
                        str(exc),
                        exc_info=True,
                    )
            for dv_def in sheet_def.data_validations:
                self.add_data_validation(ws, dv_def)
            for merge_def in sheet_def.merged_cells:
                self.merge_cells(ws, merge_def)
        return total_rows, total_charts

    def save_and_verify(self, wb: Workbook, file_path: Path, filename: str) -> int:
        """Sauvegarde le classeur et verifie la taille.

        Returns:
            Taille du fichier en octets.

        Raises:
            ExcelOutputTooLargeError: Si le fichier depasse la taille max.
        """
        wb.save(str(file_path))
        wb.close()
        file_size = file_path.stat().st_size
        size_mb = file_size / (1024 * 1024)
        if size_mb > self._max_output_size_mb:
            file_path.unlink()
            raise ExcelOutputTooLargeError(
                filename=filename,
                size_mb=size_mb,
                max_size_mb=self._max_output_size_mb,
            )
        return file_size

    def _create_sheet(self, wb: Workbook, sheet_def: SheetDef) -> Worksheet:
        """Cree une feuille avec donnees, headers, widths."""
        ws = wb.create_sheet(title=sheet_def.name)

        current_row = 1

        # Headers
        if sheet_def.headers:
            for col_idx, header in enumerate(sheet_def.headers, start=1):
                cell = ws.cell(row=current_row, column=col_idx, value=header)
                cell.font = Font(bold=True)
            current_row += 1

        # Donnees
        for row_data in sheet_def.rows:
            for col_idx, value in enumerate(row_data, start=1):
                ws.cell(row=current_row, column=col_idx, value=value)
            current_row += 1

        # Largeurs de colonnes
        if sheet_def.column_widths:
            for col_letter, width in sheet_def.column_widths.items():
                ws.column_dimensions[col_letter].width = width

        # Auto-filtre
        if sheet_def.auto_filter and sheet_def.headers:
            last_col = get_column_letter(len(sheet_def.headers))
            last_row = 1 + len(sheet_def.rows)
            ws.auto_filter.ref = f"A1:{last_col}{last_row}"

        # Gel de volets
        if sheet_def.freeze_panes:
            ws.freeze_panes = sheet_def.freeze_panes

        # Couleur d'onglet
        if sheet_def.tab_color:
            ws.sheet_properties.tabColor = sheet_def.tab_color

        return ws

    def apply_styles(self, ws: Worksheet, styles: list[CellStyleDef]) -> None:
        """Applique les styles aux plages de cellules."""
        for style_def in styles:
            font = self._build_font(style_def)
            fill = self._build_fill(style_def)
            alignment = self._build_alignment(style_def)
            border = self._build_border(style_def)

            cells_range = ws[style_def.range]
            # openpyxl retourne un Cell pour une cellule unique, sinon tuple de tuples
            if not isinstance(cells_range, tuple):
                cells_range = ((cells_range,),)

            for row in cells_range:
                for cell in row:
                    if font:
                        cell.font = font
                    if fill:
                        cell.fill = fill
                    if style_def.number_format:
                        cell.number_format = style_def.number_format
                    if alignment:
                        cell.alignment = alignment
                    if border:
                        cell.border = border

    def _build_font(self, style_def: CellStyleDef) -> Font | None:
        """Construit un objet Font a partir de la definition de style."""
        kwargs: dict[str, object] = {}
        if style_def.bold:
            kwargs["bold"] = True
        if style_def.italic:
            kwargs["italic"] = True
        if style_def.font_size:
            kwargs["size"] = style_def.font_size
        if style_def.font_color:
            kwargs["color"] = style_def.font_color
        if not kwargs:
            return None
        return Font(**kwargs)

    def _build_fill(self, style_def: CellStyleDef) -> PatternFill | None:
        """Construit un objet Fill a partir de la definition de style."""
        if not style_def.bg_color:
            return None
        return PatternFill(start_color=style_def.bg_color, fill_type="solid")

    def _build_alignment(self, style_def: CellStyleDef) -> Alignment | None:
        """Construit un objet Alignment a partir de la definition de style."""
        if not style_def.alignment and not style_def.wrap_text:
            return None
        return Alignment(
            horizontal=style_def.alignment,
            wrap_text=style_def.wrap_text,
        )

    def _build_border(self, style_def: CellStyleDef) -> Border | None:
        """Construit un objet Border a partir de la definition de style."""
        if not style_def.border:
            return None
        thin_side = Side(style="thin")
        return Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

    def add_chart(self, ws: Worksheet, chart_def: ChartDef) -> None:
        """Ajoute un graphique a la feuille.

        Raises:
            ExcelChartError: En cas d'echec de creation du graphique.
        """
        try:
            chart = self._create_chart_object(chart_def)
            if chart_def.title:
                chart.title = chart_def.title
            if chart_def.y_axis_title:
                chart.y_axis.title = chart_def.y_axis_title
            if chart_def.x_axis_title:
                chart.x_axis.title = chart_def.x_axis_title
            if chart_def.style:
                chart.style = chart_def.style

            chart.width = chart_def.width
            chart.height = chart_def.height

            sheet_name = f"'{ws.title}'" if " " in ws.title else ws.title
            data = Reference(ws, range_string=f"{sheet_name}!{chart_def.data_range}")
            chart.add_data(data, titles_from_data=True)

            if chart_def.categories_range:
                cats = Reference(ws, range_string=f"{sheet_name}!{chart_def.categories_range}")
                chart.set_categories(cats)

            ws.add_chart(chart, chart_def.position)
        except ExcelChartError:
            raise
        except Exception as exc:
            raise ExcelChartError(
                chart_title=chart_def.title,
                reason=str(exc),
            ) from exc

    def _create_chart_object(
        self, chart_def: ChartDef
    ) -> BarChart | LineChart | PieChart | ScatterChart | AreaChart:
        """Cree l'objet graphique openpyxl selon le type."""
        chart_map: dict[str, type[BarChart | LineChart | PieChart | ScatterChart | AreaChart]] = {
            "bar": BarChart,
            "line": LineChart,
            "pie": PieChart,
            "scatter": ScatterChart,
            "area": AreaChart,
            "column": BarChart,
        }
        chart_class = chart_map.get(chart_def.type)
        if not chart_class:
            msg = f"Type de graphique non supporte: {chart_def.type}"
            raise ExcelGenerationError(msg)

        chart = chart_class()
        if chart_def.type == "column" and isinstance(chart, BarChart):
            chart.type = "col"
        return chart

    def add_data_validation(self, ws: Worksheet, dv_def: DataValidationDef) -> None:
        """Ajoute une validation de donnees."""
        formula1: str | None = dv_def.formula1
        if dv_def.type == "list" and dv_def.values:
            formula1 = '"' + ",".join(dv_def.values) + '"'

        dv = DataValidation(
            type=dv_def.type,
            operator=dv_def.operator,
            formula1=formula1,
            formula2=dv_def.formula2,
        )
        if dv_def.error_message:
            dv.error = dv_def.error_message
            dv.showErrorMessage = True
        if dv_def.prompt_message:
            dv.prompt = dv_def.prompt_message
            dv.showInputMessage = True

        dv.sqref = dv_def.range
        ws.add_data_validation(dv)

    def merge_cells(self, ws: Worksheet, merge_def: MergedCellDef) -> None:
        """Fusionne des cellules."""
        ws.merge_cells(merge_def.range)
        if merge_def.value is not None:
            # Ecrire la valeur dans la premiere cellule de la plage
            top_left = merge_def.range.split(":")[0]
            ws[top_left] = merge_def.value
            ws[top_left].alignment = Alignment(horizontal="center", vertical="center")

    def sanitize_filename(self, filename: str) -> str:
        """Securise le nom de fichier (path traversal prevention).

        Args:
            filename: Nom de fichier a securiser.

        Returns:
            Nom de fichier securise.

        Raises:
            ExcelGenerationError: Si le nom est invalide ou dangereux.
        """
        if ".." in filename or "/" in filename or "\\" in filename:
            raise ExcelGenerationError(
                message=f"Nom de fichier invalide: {filename}",
                code="INVALID_FILENAME",
                suggestion="Le nom de fichier ne doit pas contenir '..' , '/' ou '\\'.",
                parameter="filename",
                retry=True,
            )

        resolved = (self._output_path / filename).resolve()
        if not resolved.is_relative_to(self._output_path.resolve()):
            raise ExcelGenerationError(
                message=f"Path traversal detecte: {filename}",
                code="PATH_TRAVERSAL",
                suggestion="Utiliser un nom de fichier simple sans chemin.",
                parameter="filename",
                retry=True,
            )

        return filename

    def ensure_output_dir(self) -> None:
        """Cree le repertoire OUTPUT_PATH s'il n'existe pas."""
        self._output_path.mkdir(parents=True, exist_ok=True)
