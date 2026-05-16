# SPDX-License-Identifier: AGPL-3.0-or-later
# Copyright (C) 2025-2026 Assistance Micro Design
"""Service d'inspection de fichiers Excel generes."""

from __future__ import annotations

import asyncio
import logging
from pathlib import Path
from typing import Any

from src.core.config import settings
from src.core.file_validation import validate_decompressed_size, validate_filename_safety


logger = logging.getLogger(__name__)

_MAX_FORMULAS = 50

_SUPPORTED_EXTENSIONS = {".xlsx"}


class FileInspector:
    """Inspecte la structure de fichiers Excel generes.

    Retourne la structure dans le vocabulaire des tools d'edition
    pour que le LLM puisse construire directement ses operations.

    Attributes:
        _output_path: Repertoire contenant les fichiers generes.
    """

    def __init__(self, output_path: Path | None = None) -> None:
        """Initialise l'inspecteur.

        Args:
            output_path: Repertoire des fichiers (defaut: settings.OUTPUT_PATH).
        """
        self._output_path = Path(output_path or settings.OUTPUT_PATH)

    async def inspect(self, filename: str, max_rows: int = 10) -> dict[str, Any]:
        """Inspecte un fichier et retourne sa structure.

        Args:
            filename: Nom du fichier (.xlsx).
            max_rows: Nombre max de lignes a afficher par feuille Excel.

        Returns:
            Structure du fichier ou dict avec "error" en cas de probleme.
        """
        # Validation anti-traversal
        if not validate_filename_safety(filename):
            return {"error": "Nom de fichier invalide", "filename": filename}

        file_path = (self._output_path / filename).resolve()
        if not file_path.is_relative_to(self._output_path.resolve()):
            return {"error": "Acces refuse (traversal)", "filename": filename}

        if not file_path.exists():
            return {"error": f"Fichier introuvable: {filename}", "filename": filename}

        ext = file_path.suffix.lower()
        if ext not in _SUPPORTED_EXTENSIONS:
            return {
                "error": f"Extension non supportee: {ext}. Seul .xlsx est accepte.",
                "filename": filename,
            }

        return await self._inspect_excel(file_path, max_rows)

    async def _inspect_excel(self, file_path: Path, max_rows: int) -> dict[str, Any]:
        """Inspecte un .xlsx via openpyxl (asyncio.to_thread)."""
        return await asyncio.to_thread(self._inspect_excel_sync, file_path, max_rows)

    def _inspect_excel_sync(self, file_path: Path, max_rows: int) -> dict[str, Any]:
        """Inspection synchrone d'un fichier Excel."""
        from openpyxl import load_workbook  # noqa: PLC0415

        try:
            validate_decompressed_size(file_path, settings.MAX_DECOMPRESSED_MB)
            wb_formulas = load_workbook(file_path, data_only=False)
            wb_values = load_workbook(file_path, data_only=True)
        except Exception as exc:
            return {"error": f"Fichier illisible: {exc}", "filename": file_path.name}

        sheets = []
        for ws_f, ws_v in zip(wb_formulas.worksheets, wb_values.worksheets, strict=True):
            sheet_info = self._extract_sheet_info(ws_f, ws_v, max_rows)
            sheets.append(sheet_info)

        wb_formulas.close()
        wb_values.close()

        return {
            "filename": file_path.name,
            "type": "excel",
            "file_size_bytes": file_path.stat().st_size,
            "editable_with": "edit_excel_document",
            "sheets": sheets,
        }

    def _extract_sheet_info(self, ws_f: Any, ws_v: Any, max_rows: int) -> dict[str, Any]:
        """Extrait les informations d'une feuille Excel.

        Args:
            ws_f: Worksheet avec formules (data_only=False).
            ws_v: Worksheet avec valeurs calculees (data_only=True).
            max_rows: Nombre max de lignes d'echantillon.

        Returns:
            Dictionnaire avec structure, donnees et proprietes de la feuille.
        """
        sheet_info: dict[str, Any] = {
            "name": ws_f.title,
            "dimensions": ws_f.dimensions,
            "row_count": (ws_f.max_row - 1) if ws_f.max_row else 0,
            "column_count": ws_f.max_column or 0,
        }

        headers: list[Any] = []
        if ws_v.max_row:
            headers = [cell.value for cell in ws_v[1] if cell.value is not None]
        sheet_info["headers"] = headers

        sheet_info["sample_data"] = self._extract_sample_data(ws_v, max_rows)

        formulas, total_formulas = self._extract_formulas(ws_f, ws_v)
        sheet_info["formulas"] = formulas
        sheet_info["total_formulas"] = total_formulas
        if total_formulas > _MAX_FORMULAS:
            sheet_info["formulas_truncated"] = True

        sheet_info["charts"] = self._extract_excel_charts(ws_f)
        sheet_info["merged_cells"] = [str(r) for r in ws_f.merged_cells.ranges]
        sheet_info["data_validations"] = self._extract_data_validations(ws_f)
        sheet_info.update(self._extract_sheet_properties(ws_f))

        return sheet_info

    def _extract_sample_data(self, ws_v: Any, max_rows: int) -> dict[str, Any]:
        """Extrait les donnees d'echantillon (format {cellule: valeur}).

        Args:
            ws_v: Worksheet avec valeurs calculees.
            max_rows: Nombre max de lignes a extraire.

        Returns:
            Dictionnaire {coordonnee: valeur} des cellules non vides.
        """
        sample: dict[str, Any] = {}
        for row_idx in range(2, min(2 + max_rows, (ws_v.max_row or 1) + 1)):
            for cell in ws_v[row_idx]:
                if cell.value is not None:
                    sample[cell.coordinate] = cell.value
        return sample

    def _extract_excel_charts(self, ws_f: Any) -> list[dict[str, Any]]:
        """Extrait les graphiques d'une feuille Excel.

        Args:
            ws_f: Worksheet avec formules.

        Returns:
            Liste de dictionnaires {title, type, data_range}.
        """
        charts = []
        # openpyxl: pas d'API publique pour acceder aux graphiques
        for c in ws_f._charts:
            chart_info: dict[str, Any] = {"title": c.title or "", "type": c.type}
            if c.series:
                try:
                    chart_info["data_range"] = str(c.series[0].val.numRef)
                except (AttributeError, IndexError):
                    chart_info["data_range"] = None
            charts.append(chart_info)
        return charts

    def _extract_data_validations(self, ws_f: Any) -> list[dict[str, Any]]:
        """Extrait les regles de validation de donnees.

        Args:
            ws_f: Worksheet avec formules.

        Returns:
            Liste de dictionnaires {range, type, formula1}.
        """
        return [
            {"range": str(dv.sqref), "type": dv.type or "", "formula1": dv.formula1}
            for dv in (ws_f.data_validations.dataValidation if ws_f.data_validations else [])
        ]

    def _extract_sheet_properties(self, ws_f: Any) -> dict[str, Any]:
        """Extrait les proprietes visuelles de la feuille.

        Args:
            ws_f: Worksheet avec formules.

        Returns:
            Dictionnaire avec column_widths, freeze_panes, auto_filter, tab_color.
        """
        widths: dict[str, float] = {}
        for col_letter, dim in ws_f.column_dimensions.items():
            if dim.width and dim.width != 8.43:
                widths[col_letter] = round(dim.width, 1)
        return {
            "column_widths": widths,
            "freeze_panes": str(ws_f.freeze_panes) if ws_f.freeze_panes else None,
            "auto_filter": ws_f.auto_filter.ref is not None if ws_f.auto_filter else False,
            "tab_color": (
                ws_f.sheet_properties.tabColor.rgb if ws_f.sheet_properties.tabColor else None
            ),
        }

    def _extract_formulas(self, ws_f: Any, ws_v: Any) -> tuple[list[dict[str, Any]], int]:
        """Extrait les formules d'une feuille (limitees a _MAX_FORMULAS).

        Args:
            ws_f: Worksheet avec formules (data_only=False).
            ws_v: Worksheet avec valeurs calculees (data_only=True).

        Returns:
            Tuple (liste de formules, nombre total de formules).
        """
        formulas: list[dict[str, Any]] = []
        total_formulas = 0
        for row in ws_f.iter_rows():
            for cell in row:
                if not isinstance(cell.value, str):
                    continue
                if not cell.value.startswith("="):
                    continue
                total_formulas += 1
                if len(formulas) < _MAX_FORMULAS:
                    val_cell = ws_v[cell.coordinate]
                    formulas.append(
                        {
                            "cell": cell.coordinate,
                            "formula": cell.value,
                            "result": val_cell.value,
                        }
                    )
        return formulas, total_formulas
