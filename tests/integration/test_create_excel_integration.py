# SPDX-License-Identifier: AGPL-3.0-or-later
# Copyright (C) 2026 Assistance Micro Design
"""Tests d'integration pour la creation de fichiers Excel."""

from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import load_workbook

from src.mcp.tools.create_excel import CreateExcelTool
from src.services.excel.generator import ExcelGenerator


@pytest.fixture
def create_excel_tool(tmp_path: Path) -> CreateExcelTool:
    """Tool MCP pret a l'emploi avec output tmp."""
    tool = CreateExcelTool()
    tool._initialized = True
    tool._generator = ExcelGenerator(output_path=tmp_path)
    return tool


@pytest.mark.integration
class TestCreateExcelE2E:
    """Tests E2E complets de creation Excel."""

    @pytest.mark.asyncio
    async def test_full_excel_creation(self, create_excel_tool: CreateExcelTool) -> None:
        """Scenario complet: donnees, formules, styles, graphique, validation."""
        result = await create_excel_tool.execute(
            {
                "filename": "rapport-ventes-2026.xlsx",
                "sheets": [
                    {
                        "name": "Ventes Q1",
                        "headers": ["Produit", "Janvier", "Fevrier", "Mars", "Total"],
                        "rows": [
                            ["Widget A", 100, 150, 200, "=SUM(B2:D2)"],
                            ["Widget B", 80, 120, 160, "=SUM(B3:D3)"],
                            [
                                "Total",
                                "=SUM(B2:B3)",
                                "=SUM(C2:C3)",
                                "=SUM(D2:D3)",
                                "=SUM(E2:E3)",
                            ],
                        ],
                        "column_widths": {
                            "A": 20,
                            "B": 12,
                            "C": 12,
                            "D": 12,
                            "E": 15,
                        },
                        "styles": [
                            {
                                "range": "A1:E1",
                                "bold": True,
                                "bg_color": "4472C4",
                                "font_color": "FFFFFF",
                            },
                            {
                                "range": "E2:E4",
                                "number_format": "#,##0",
                                "bold": True,
                            },
                        ],
                        "auto_filter": True,
                        "freeze_panes": "A2",
                        "charts": [
                            {
                                "type": "bar",
                                "title": "Ventes par mois",
                                "data_range": "B1:D3",
                                "categories_range": "A2:A3",
                                "position": "G2",
                                "y_axis_title": "Montant (EUR)",
                            }
                        ],
                        "data_validations": [
                            {
                                "range": "A5:A100",
                                "type": "list",
                                "values": ["Widget A", "Widget B", "Widget C"],
                            }
                        ],
                    }
                ],
                "author": "MCP Zileo RAG",
            }
        )

        # Verifier le resultat
        assert result["filename"] == "rapport-ventes-2026.xlsx"
        assert result["sheets_created"] == 1
        assert result["total_rows"] == 3
        assert result["total_charts"] == 1
        assert result["file_size_bytes"] > 0
        assert result["overwritten"] is False

        # Verifier le fichier est lisible
        file_path = result["file_path"]
        assert Path(file_path).exists()

        wb = load_workbook(file_path)
        ws = wb["Ventes Q1"]

        # Verifier les donnees
        assert ws["A1"].value == "Produit"
        assert ws["B2"].value == 100
        assert ws["E2"].value == "=SUM(B2:D2)"

        # Verifier le style bold sur header
        assert ws["A1"].font.bold is True

        # Verifier le gel
        assert ws.freeze_panes == "A2"

        # Verifier auto-filter
        assert ws.auto_filter.ref is not None

        # Verifier les largeurs
        assert ws.column_dimensions["A"].width == 20

        # Verifier le graphique
        assert len(ws._charts) == 1

        # Verifier la validation
        assert len(ws.data_validations.dataValidation) == 1

        # Verifier l'auteur
        assert wb.properties.creator == "MCP Zileo RAG"

        wb.close()

    @pytest.mark.asyncio
    async def test_overwrite_returns_flag(self, create_excel_tool: CreateExcelTool) -> None:
        """Overwrite d'un fichier existant retourne overwritten=True."""
        args = {
            "filename": "overwrite.xlsx",
            "sheets": [{"name": "S1", "rows": [["data"]]}],
        }

        r1 = await create_excel_tool.execute(args)
        assert r1["overwritten"] is False

        r2 = await create_excel_tool.execute(args)
        assert r2["overwritten"] is True
