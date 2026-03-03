# SPDX-License-Identifier: AGPL-3.0-or-later
# Copyright (C) 2026 Assistance Micro Design
"""Tests d'integration pour edit_excel_document (create -> edit -> verify)."""

from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import load_workbook

from src.models.api import CreateExcelParams, EditExcelParams
from src.models.excel_edit import (
    AddChartOp,
    AddSheetOp,
    ApplyStylesOp,
    DeleteSheetOp,
    InsertRowsOp,
    UpdateCellsOp,
)
from src.models.excel_generation import CellStyleDef, ChartDef, SheetDef
from src.services.excel.editor import ExcelEditor
from src.services.excel.generator import ExcelGenerator


@pytest.mark.integration
class TestCreateThenEdit:
    """Integration: creer un fichier puis le modifier."""

    @pytest.mark.asyncio
    async def test_full_round_trip(self, tmp_path: Path) -> None:
        """Create -> Edit (multi-ops) -> Verify le fichier final."""
        gen = ExcelGenerator(output_path=tmp_path)
        editor = ExcelEditor(output_path=tmp_path)

        # 1. Creer le fichier
        create_result = await gen.generate(
            CreateExcelParams(
                filename="rapport.xlsx",
                sheets=[
                    SheetDef(
                        name="Donnees",
                        headers=["Mois", "CA", "Charges"],
                        rows=[
                            ["Jan", 10000, 8000],
                            ["Fev", 12000, 8500],
                            ["Mar", 15000, 9000],
                        ],
                    ),
                    SheetDef(name="Temp", rows=[["placeholder"]]),
                ],
            )
        )
        assert create_result.sheets_created == 2

        # 2. Editer le fichier avec plusieurs operations
        edit_result = await editor.edit(
            EditExcelParams(
                filename="rapport.xlsx",
                operations=[
                    # Corriger une valeur
                    UpdateCellsOp(sheet="Donnees", cells={"B2": 11000}),
                    # Ajouter une ligne
                    InsertRowsOp(sheet="Donnees", rows=[["Avr", 18000, 10000]]),
                    # Ajouter une formule
                    UpdateCellsOp(sheet="Donnees", cells={"D2": "=B2-C2"}),
                    # Styler les headers
                    ApplyStylesOp(
                        sheet="Donnees",
                        styles=[CellStyleDef(range="A1:D1", bold=True, bg_color="4472C4")],
                    ),
                    # Ajouter un graphique
                    AddChartOp(
                        sheet="Donnees",
                        chart=ChartDef(
                            type="bar",
                            data_range="B1:C5",
                            categories_range="A2:A5",
                            title="CA vs Charges",
                            position="F1",
                        ),
                    ),
                    # Supprimer la feuille temporaire
                    DeleteSheetOp(name="Temp"),
                    # Ajouter une feuille de synthese
                    AddSheetOp(
                        name="Synthese",
                        headers=["Metrique", "Valeur"],
                        rows=[["Total CA", "=SUM(Donnees!B2:B5)"]],
                    ),
                ],
            )
        )

        assert edit_result.operations_applied == 7
        assert edit_result.operations_skipped == 0

        # 3. Verifier le fichier final
        wb = load_workbook(edit_result.file_path)

        # Feuilles correctes
        assert "Donnees" in wb.sheetnames
        assert "Synthese" in wb.sheetnames
        assert "Temp" not in wb.sheetnames

        # Valeur corrigee
        assert wb["Donnees"]["B2"].value == 11000

        # Ligne ajoutee
        assert wb["Donnees"]["A5"].value == "Avr"

        # Formule ecrite
        assert wb["Donnees"]["D2"].value == "=B2-C2"

        # Style applique
        assert wb["Donnees"]["A1"].font.bold is True

        # Graphique present
        assert len(wb["Donnees"]._charts) == 1

        # Synthese creee
        assert wb["Synthese"]["A1"].value == "Metrique"

        wb.close()
