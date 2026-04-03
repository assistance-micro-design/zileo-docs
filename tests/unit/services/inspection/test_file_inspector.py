"""Tests unitaires pour FileInspector service."""

from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import Workbook

from src.services.inspection.file_inspector import FileInspector


@pytest.fixture
def inspector(tmp_path: Path) -> FileInspector:
    """FileInspector avec output_path temporaire."""
    return FileInspector(output_path=tmp_path)


@pytest.fixture
def sample_xlsx(tmp_path: Path) -> Path:
    """Cree un fichier Excel de test."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Ventes"
    ws.append(["Nom", "CA Q1", "CA Q2", "Total"])
    ws.append(["Alice", 15000, 18000, "=SUM(B2:C2)"])
    ws.append(["Bob", 12000, 14000, "=SUM(B3:C3)"])
    ws.merge_cells("A5:B5")
    ws.column_dimensions["A"].width = 20.0
    ws.freeze_panes = "A2"
    path = tmp_path / "rapport.xlsx"
    wb.save(path)
    return path


@pytest.fixture
def empty_xlsx(tmp_path: Path) -> Path:
    """Cree un fichier Excel vide."""
    wb = Workbook()
    path = tmp_path / "empty.xlsx"
    wb.save(path)
    return path


class TestInspectExcel:
    """Tests pour l'inspection de fichiers Excel."""

    @pytest.mark.asyncio
    async def test_inspect_excel_returns_sheets_structure(
        self, inspector: FileInspector, sample_xlsx: Path
    ) -> None:
        result = await inspector.inspect(sample_xlsx.name)

        assert result["type"] == "excel"
        assert "sheets" in result
        assert len(result["sheets"]) == 1
        assert result["sheets"][0]["name"] == "Ventes"

    @pytest.mark.asyncio
    async def test_inspect_excel_headers_extracted(
        self, inspector: FileInspector, sample_xlsx: Path
    ) -> None:
        result = await inspector.inspect(sample_xlsx.name)

        sheet = result["sheets"][0]
        assert sheet["headers"] == ["Nom", "CA Q1", "CA Q2", "Total"]

    @pytest.mark.asyncio
    async def test_inspect_excel_sample_data_format_matches_update_cells(
        self, inspector: FileInspector, sample_xlsx: Path
    ) -> None:
        result = await inspector.inspect(sample_xlsx.name)

        sample = result["sheets"][0]["sample_data"]
        assert isinstance(sample, dict)
        assert sample["A2"] == "Alice"
        assert sample["B2"] == 15000

    @pytest.mark.asyncio
    async def test_inspect_excel_formulas_extracted(
        self, inspector: FileInspector, sample_xlsx: Path
    ) -> None:
        result = await inspector.inspect(sample_xlsx.name)

        formulas = result["sheets"][0]["formulas"]
        assert len(formulas) == 2
        assert formulas[0]["cell"] == "D2"
        assert formulas[0]["formula"] == "=SUM(B2:C2)"

    @pytest.mark.asyncio
    async def test_inspect_excel_merged_cells_extracted(
        self, inspector: FileInspector, sample_xlsx: Path
    ) -> None:
        result = await inspector.inspect(sample_xlsx.name)

        merged = result["sheets"][0]["merged_cells"]
        assert "A5:B5" in merged

    @pytest.mark.asyncio
    async def test_inspect_excel_column_widths_extracted(
        self, inspector: FileInspector, sample_xlsx: Path
    ) -> None:
        result = await inspector.inspect(sample_xlsx.name)

        widths = result["sheets"][0]["column_widths"]
        assert "A" in widths
        assert widths["A"] == 20.0

    @pytest.mark.asyncio
    async def test_inspect_excel_max_rows_limits_sample(
        self, inspector: FileInspector, sample_xlsx: Path
    ) -> None:
        result = await inspector.inspect(sample_xlsx.name, max_rows=1)

        sample = result["sheets"][0]["sample_data"]
        # Only row 2 should be present (max_rows=1)
        assert "A2" in sample
        assert "A3" not in sample

    @pytest.mark.asyncio
    async def test_inspect_excel_empty_sheet(
        self, inspector: FileInspector, empty_xlsx: Path
    ) -> None:
        result = await inspector.inspect(empty_xlsx.name)

        sheet = result["sheets"][0]
        assert sheet["headers"] == []
        assert sheet["sample_data"] == {}
        assert sheet["formulas"] == []

    @pytest.mark.asyncio
    async def test_inspect_excel_editable_with_field(
        self, inspector: FileInspector, sample_xlsx: Path
    ) -> None:
        result = await inspector.inspect(sample_xlsx.name)

        assert result["editable_with"] == "edit_excel_document"

    @pytest.mark.asyncio
    async def test_inspect_excel_freeze_panes(
        self, inspector: FileInspector, sample_xlsx: Path
    ) -> None:
        result = await inspector.inspect(sample_xlsx.name)

        assert result["sheets"][0]["freeze_panes"] == "A2"


class TestInspectErrors:
    """Tests pour les cas d'erreur."""

    @pytest.mark.asyncio
    async def test_unsupported_extension(self, inspector: FileInspector, tmp_path: Path) -> None:
        (tmp_path / "file.txt").write_text("not supported")

        result = await inspector.inspect("file.txt")

        assert "error" in result

    @pytest.mark.asyncio
    async def test_file_not_found(self, inspector: FileInspector) -> None:
        result = await inspector.inspect("nonexistent.xlsx")

        assert "error" in result

    @pytest.mark.asyncio
    async def test_path_traversal_blocked(self, inspector: FileInspector) -> None:
        result = await inspector.inspect("../../../etc/passwd")

        assert "error" in result

    @pytest.mark.asyncio
    async def test_corrupted_xlsx(self, inspector: FileInspector, tmp_path: Path) -> None:
        (tmp_path / "corrupt.xlsx").write_text("not a real xlsx")

        result = await inspector.inspect("corrupt.xlsx")

        assert "error" in result


class TestInspectExcelFormulasLimit:
    """Tests pour la limite de formules."""

    @pytest.mark.asyncio
    async def test_excel_formulas_limited_to_50(self, tmp_path: Path) -> None:
        wb = Workbook()
        ws = wb.active
        ws.append(["Header"])
        for i in range(2, 62):
            ws.append([f"=SUM(A{i - 1})"])
        path = tmp_path / "many_formulas.xlsx"
        wb.save(path)

        inspector = FileInspector(output_path=tmp_path)
        result = await inspector.inspect("many_formulas.xlsx")

        sheet = result["sheets"][0]
        assert len(sheet["formulas"]) == 50
        assert sheet["total_formulas"] == 60
        assert sheet["formulas_truncated"] is True
