"""Tests pour l'extracteur Excel."""

from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import Workbook

from src.models.excel import (
    CellType,
    ExcelCell,
    ExcelDocument,
    ExcelFormula,
    ExcelSheet,
    ExcelTable,
)
from src.services.excel.extractor import ExcelExtractor


class TestExcelExtractor:
    """Tests pour l'extracteur Excel."""

    @pytest.fixture
    def extractor(self) -> ExcelExtractor:
        """Create an ExcelExtractor instance."""
        return ExcelExtractor()

    @pytest.fixture
    def sample_xlsx(self, tmp_path: Path) -> Path:
        """Crée un fichier Excel de test."""
        wb = Workbook()
        ws = wb.active
        if ws is None:
            ws = wb.create_sheet("Données")
        else:
            ws.title = "Données"

        # En-têtes
        ws["A1"] = "Produit"
        ws["B1"] = "Quantité"
        ws["C1"] = "Prix"
        ws["D1"] = "Total"

        # Données
        ws["A2"] = "Widget A"
        ws["B2"] = 10
        ws["C2"] = 25.50
        ws["D2"] = "=B2*C2"  # Formule

        ws["A3"] = "Widget B"
        ws["B3"] = 5
        ws["C3"] = 15.00
        ws["D3"] = "=B3*C3"

        # Total
        ws["D4"] = "=SUM(D2:D3)"

        file_path = tmp_path / "test.xlsx"
        wb.save(file_path)
        wb.close()

        return file_path

    @pytest.fixture
    def empty_xlsx(self, tmp_path: Path) -> Path:
        """Crée un fichier Excel vide."""
        wb = Workbook()
        file_path = tmp_path / "empty.xlsx"
        wb.save(file_path)
        wb.close()
        return file_path

    @pytest.fixture
    def multi_sheet_xlsx(self, tmp_path: Path) -> Path:
        """Crée un fichier Excel avec plusieurs feuilles."""
        wb = Workbook()

        # Feuille 1
        ws1 = wb.active
        if ws1 is None:
            ws1 = wb.create_sheet("Feuille1")
        else:
            ws1.title = "Feuille1"
        ws1["A1"] = "Données Feuille 1"

        # Feuille 2
        ws2 = wb.create_sheet("Feuille2")
        ws2["A1"] = "Données Feuille 2"
        ws2["B1"] = 42

        # Feuille 3
        ws3 = wb.create_sheet("Calculs")
        ws3["A1"] = "=Feuille2!B1*2"

        file_path = tmp_path / "multi_sheet.xlsx"
        wb.save(file_path)
        wb.close()

        return file_path

    async def test_extract_xlsx_basic(
        self,
        extractor: ExcelExtractor,
        sample_xlsx: Path,
    ) -> None:
        """Test extraction basique d'un fichier xlsx."""
        doc = await extractor.extract(sample_xlsx)

        assert isinstance(doc, ExcelDocument)
        assert doc.format == "xlsx"
        assert len(doc.sheets) == 1
        assert doc.sheets[0].name == "Données"

    async def test_extract_formulas(
        self,
        extractor: ExcelExtractor,
        sample_xlsx: Path,
    ) -> None:
        """Test que les formules sont extraites."""
        doc = await extractor.extract(sample_xlsx)

        formulas = doc.get_all_formulas()
        assert len(formulas) == 3

        # Vérifier la formule SUM
        sum_formula = next((f for f in formulas if "SUM" in f.formula), None)
        assert sum_formula is not None
        assert sum_formula.formula == "=SUM(D2:D3)"
        assert sum_formula.cell == "D4"

    async def test_extract_cell_types(
        self,
        extractor: ExcelExtractor,
        sample_xlsx: Path,
    ) -> None:
        """Test que les types de cellules sont correctement détectés."""
        doc = await extractor.extract(sample_xlsx)

        sheet = doc.sheets[0]

        # A1 is text
        assert sheet.cells[0][0].cell_type == CellType.TEXT

        # B2 is number
        assert sheet.cells[1][1].cell_type == CellType.NUMBER

        # D2 is formula
        assert sheet.cells[1][3].cell_type == CellType.FORMULA

    async def test_extract_empty_xlsx(
        self,
        extractor: ExcelExtractor,
        empty_xlsx: Path,
    ) -> None:
        """Test extraction d'un fichier xlsx vide."""
        doc = await extractor.extract(empty_xlsx)

        assert isinstance(doc, ExcelDocument)
        assert len(doc.sheets) == 1
        assert doc.total_formulas == 0

    async def test_extract_multi_sheet(
        self,
        extractor: ExcelExtractor,
        multi_sheet_xlsx: Path,
    ) -> None:
        """Test extraction de plusieurs feuilles."""
        doc = await extractor.extract(multi_sheet_xlsx)

        assert len(doc.sheets) == 3
        sheet_names = [s.name for s in doc.sheets]
        assert "Feuille1" in sheet_names
        assert "Feuille2" in sheet_names
        assert "Calculs" in sheet_names

    async def test_file_not_found(self, extractor: ExcelExtractor) -> None:
        """Test erreur fichier non trouvé."""
        with pytest.raises(FileNotFoundError):
            await extractor.extract("/non/existent/file.xlsx")

    async def test_unsupported_format(
        self,
        extractor: ExcelExtractor,
        tmp_path: Path,
    ) -> None:
        """Test erreur format non supporté."""
        bad_file = tmp_path / "test.txt"
        bad_file.write_text("not excel")

        with pytest.raises(ValueError, match="Format non supporté"):
            await extractor.extract(bad_file)

    async def test_to_markdown(
        self,
        extractor: ExcelExtractor,
        sample_xlsx: Path,
    ) -> None:
        """Test conversion en Markdown."""
        doc = await extractor.extract(sample_xlsx)

        md = doc.to_markdown()

        assert "test.xlsx" in md
        assert "Feuille: Données" in md

    async def test_formula_dependencies(
        self,
        extractor: ExcelExtractor,
        sample_xlsx: Path,
    ) -> None:
        """Test que les dépendances des formules sont parsées."""
        doc = await extractor.extract(sample_xlsx)

        formulas = doc.get_all_formulas()
        sum_formula = next((f for f in formulas if "SUM" in f.formula), None)

        assert sum_formula is not None
        assert "D2" in sum_formula.dependencies
        assert "D3" in sum_formula.dependencies


class TestExcelFormulaParsing:
    """Tests pour le parsing des formules."""

    def test_parse_simple_reference(self) -> None:
        """Test parsing de références simples."""
        extractor = ExcelExtractor()
        deps = extractor._parse_formula_dependencies("=A1+B1")
        assert set(deps) == {"A1", "B1"}

    def test_parse_range_reference(self) -> None:
        """Test parsing de références de plage."""
        extractor = ExcelExtractor()
        deps = extractor._parse_formula_dependencies("=SUM(A1:A10)")
        assert "A1" in deps
        assert "A10" in deps

    def test_parse_absolute_reference(self) -> None:
        """Test parsing de références absolues."""
        extractor = ExcelExtractor()
        deps = extractor._parse_formula_dependencies("=$A$1+B2")
        assert "$A$1" in deps
        assert "B2" in deps

    def test_parse_mixed_references(self) -> None:
        """Test parsing de références mixtes."""
        extractor = ExcelExtractor()
        deps = extractor._parse_formula_dependencies("=A$1+$B2")
        assert "A$1" in deps
        assert "$B2" in deps

    def test_parse_complex_formula(self) -> None:
        """Test parsing d'une formule complexe."""
        extractor = ExcelExtractor()
        deps = extractor._parse_formula_dependencies("=IF(A1>10,SUM(B1:B5),AVERAGE(C1:C10))")
        assert "A1" in deps
        assert "B1" in deps
        assert "B5" in deps
        assert "C1" in deps
        assert "C10" in deps


class TestExcelModels:
    """Tests pour les modèles Excel."""

    def test_excel_cell_to_dict(self) -> None:
        """Test conversion ExcelCell en dict."""
        cell = ExcelCell(
            row=1,
            column=1,
            column_letter="A",
            value="test",
            formula=None,
            cell_type=CellType.TEXT,
        )

        d = cell.to_dict()

        assert d["cell"] == "A1"
        assert d["value"] == "test"
        assert d["type"] == "text"

    def test_excel_table_to_markdown(self) -> None:
        """Test conversion ExcelTable en Markdown."""
        table = ExcelTable(
            name="TestTable",
            range="A1:C3",
            headers=["A", "B", "C"],
            data=[
                [1, 2, 3],
                [4, 5, 6],
            ],
        )

        md = table.to_markdown()

        assert "| A | B | C |" in md
        assert "| --- | --- | --- |" in md
        assert "| 1 | 2 | 3 |" in md
        assert "| 4 | 5 | 6 |" in md

    def test_excel_sheet_get_text_content(self) -> None:
        """Test génération du contenu textuel."""
        sheet = ExcelSheet(
            name="TestSheet",
            index=0,
            rows_count=10,
            columns_count=5,
            tables=[
                ExcelTable(
                    name="Table1",
                    range="A1:B2",
                    headers=["X", "Y"],
                    data=[[1, 2]],
                )
            ],
            formulas=[
                ExcelFormula(
                    cell="C1",
                    sheet="TestSheet",
                    formula="=A1+B1",
                    result=3,
                )
            ],
        )

        content = sheet.get_text_content()

        assert "## Feuille: TestSheet" in content
        assert "### Tableau: Table1" in content
        assert "### Formules" in content
        assert "`C1`" in content
        assert "`=A1+B1`" in content

    def test_excel_document_get_all_formulas(self) -> None:
        """Test récupération de toutes les formules."""
        doc = ExcelDocument(
            filename="test.xlsx",
            file_path="/path/test.xlsx",
            format="xlsx",
            sheets=[
                ExcelSheet(
                    name="Sheet1",
                    index=0,
                    rows_count=1,
                    columns_count=1,
                    formulas=[
                        ExcelFormula(
                            cell="A1",
                            sheet="Sheet1",
                            formula="=1+1",
                            result=2,
                        )
                    ],
                ),
                ExcelSheet(
                    name="Sheet2",
                    index=1,
                    rows_count=1,
                    columns_count=1,
                    formulas=[
                        ExcelFormula(
                            cell="B1",
                            sheet="Sheet2",
                            formula="=2+2",
                            result=4,
                        )
                    ],
                ),
            ],
        )

        formulas = doc.get_all_formulas()

        assert len(formulas) == 2
        cells = [f.cell for f in formulas]
        assert "A1" in cells
        assert "B1" in cells
