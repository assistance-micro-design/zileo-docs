# SPDX-License-Identifier: AGPL-3.0-or-later
# Copyright (C) 2026 Assistance Micro Design
"""Tests pour le service ExcelGenerator."""

from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import load_workbook

from src.core.exceptions import (
    ExcelChartError,
    ExcelFormulaInjectionError,
    ExcelGenerationError,
)
from src.models.api import CreateExcelParams, CreateExcelResult
from src.models.excel_generation import (
    CellStyleDef,
    ChartDef,
    DataValidationDef,
    MergedCellDef,
    SheetDef,
)
from src.services.excel.generator import ExcelGenerator


@pytest.fixture
def generator(tmp_path: Path) -> ExcelGenerator:
    """Generateur avec output dans tmp_path."""
    return ExcelGenerator(output_path=tmp_path)


@pytest.fixture
def minimal_params() -> CreateExcelParams:
    """Params minimaux valides."""
    return CreateExcelParams(
        filename="test.xlsx",
        sheets=[
            SheetDef(
                name="Feuille1",
                headers=["Nom", "Age"],
                rows=[["Alice", 30], ["Bob", 25]],
            )
        ],
    )


class TestExcelGeneratorCore:
    """Tests core du generateur."""

    @pytest.mark.asyncio
    async def test_generate_minimal(
        self, generator: ExcelGenerator, minimal_params: CreateExcelParams
    ) -> None:
        """Generation d'un classeur minimal."""
        result = await generator.generate(minimal_params)

        assert isinstance(result, CreateExcelResult)
        assert result.filename == "test.xlsx"
        assert result.sheets_created == 1
        assert result.total_rows == 2
        assert result.total_charts == 0
        assert result.file_size_bytes > 0
        assert result.overwritten is False
        assert Path(result.file_path).exists()

    @pytest.mark.asyncio
    async def test_generate_file_readable(
        self, generator: ExcelGenerator, minimal_params: CreateExcelParams
    ) -> None:
        """Le fichier genere est lisible par openpyxl."""
        result = await generator.generate(minimal_params)
        wb = load_workbook(result.file_path)
        ws = wb.active

        assert ws is not None
        assert ws.title == "Feuille1"
        # Headers en ligne 1
        assert ws["A1"].value == "Nom"
        assert ws["B1"].value == "Age"
        # Donnees en lignes 2-3
        assert ws["A2"].value == "Alice"
        assert ws["B2"].value == 30
        assert ws["A3"].value == "Bob"
        assert ws["B3"].value == 25
        wb.close()

    @pytest.mark.asyncio
    async def test_generate_multi_sheets(self, generator: ExcelGenerator) -> None:
        """Generation multi-feuilles."""
        params = CreateExcelParams(
            filename="multi.xlsx",
            sheets=[
                SheetDef(name="Ventes", rows=[["A", 1]]),
                SheetDef(name="Achats", rows=[["B", 2]]),
                SheetDef(name="Resume", rows=[["C", 3]]),
            ],
        )
        result = await generator.generate(params)

        assert result.sheets_created == 3
        wb = load_workbook(result.file_path)
        assert wb.sheetnames == ["Ventes", "Achats", "Resume"]
        wb.close()

    @pytest.mark.asyncio
    async def test_generate_with_formulas(self, generator: ExcelGenerator) -> None:
        """Les formules sont ecrites correctement."""
        params = CreateExcelParams(
            filename="formulas.xlsx",
            sheets=[
                SheetDef(
                    name="Calculs",
                    headers=["A", "B", "Total"],
                    rows=[[10, 20, "=SUM(A2:B2)"]],
                )
            ],
        )
        result = await generator.generate(params)

        wb = load_workbook(result.file_path)
        ws = wb.active
        assert ws is not None
        assert ws["C2"].value == "=SUM(A2:B2)"
        wb.close()

    @pytest.mark.asyncio
    async def test_generate_column_widths(self, generator: ExcelGenerator) -> None:
        """Largeurs de colonnes appliquees."""
        params = CreateExcelParams(
            filename="widths.xlsx",
            sheets=[
                SheetDef(
                    name="Test",
                    rows=[["data"]],
                    column_widths={"A": 30, "B": 15},
                )
            ],
        )
        result = await generator.generate(params)

        wb = load_workbook(result.file_path)
        ws = wb.active
        assert ws is not None
        assert ws.column_dimensions["A"].width == 30
        assert ws.column_dimensions["B"].width == 15
        wb.close()

    @pytest.mark.asyncio
    async def test_generate_auto_filter(self, generator: ExcelGenerator) -> None:
        """Auto-filtre applique."""
        params = CreateExcelParams(
            filename="filter.xlsx",
            sheets=[
                SheetDef(
                    name="Test",
                    headers=["Col1", "Col2"],
                    rows=[["a", "b"]],
                    auto_filter=True,
                )
            ],
        )
        result = await generator.generate(params)

        wb = load_workbook(result.file_path)
        ws = wb.active
        assert ws is not None
        assert ws.auto_filter.ref is not None
        wb.close()

    @pytest.mark.asyncio
    async def test_generate_freeze_panes(self, generator: ExcelGenerator) -> None:
        """Gel de volets applique."""
        params = CreateExcelParams(
            filename="freeze.xlsx",
            sheets=[
                SheetDef(
                    name="Test",
                    headers=["H1", "H2"],
                    rows=[["a", "b"]],
                    freeze_panes="A2",
                )
            ],
        )
        result = await generator.generate(params)

        wb = load_workbook(result.file_path)
        ws = wb.active
        assert ws is not None
        assert ws.freeze_panes == "A2"
        wb.close()

    @pytest.mark.asyncio
    async def test_generate_with_author(self, generator: ExcelGenerator) -> None:
        """Metadonnee auteur du classeur."""
        params = CreateExcelParams(
            filename="author.xlsx",
            sheets=[SheetDef(name="F1")],
            author="MCP Zileo RAG",
        )
        result = await generator.generate(params)

        wb = load_workbook(result.file_path)
        assert wb.properties.creator == "MCP Zileo RAG"
        wb.close()

    @pytest.mark.asyncio
    async def test_generate_total_rows_counts_all_sheets(self, generator: ExcelGenerator) -> None:
        """total_rows est la somme sur toutes les feuilles."""
        params = CreateExcelParams(
            filename="count.xlsx",
            sheets=[
                SheetDef(name="A", rows=[["x"], ["y"]]),
                SheetDef(name="B", rows=[["z"]]),
            ],
        )
        result = await generator.generate(params)
        assert result.total_rows == 3


class TestExcelGeneratorOverwrite:
    """Tests ecrasement de fichier existant."""

    @pytest.mark.asyncio
    async def test_overwrite_existing_file(
        self, generator: ExcelGenerator, minimal_params: CreateExcelParams
    ) -> None:
        """Ecrasement silencieux avec overwritten=True."""
        # Premier fichier
        result1 = await generator.generate(minimal_params)
        assert result1.overwritten is False

        # Deuxieme fichier (meme nom)
        result2 = await generator.generate(minimal_params)
        assert result2.overwritten is True


class TestExcelGeneratorSecurity:
    """Tests securite du generateur."""

    @pytest.mark.asyncio
    async def test_sanitize_path_traversal(self, generator: ExcelGenerator) -> None:
        """Path traversal dans filename bloque."""
        params = CreateExcelParams(
            filename="test.xlsx",
            sheets=[SheetDef(name="F1")],
        )
        # Meme si la regex Pydantic bloque .., tester le runtime aussi
        result = await generator.generate(params)
        assert Path(result.file_path).parent == generator._output_path

    def test_ensure_output_dir_creates_directory(self, tmp_path: Path) -> None:
        """ensure_output_dir cree le repertoire s'il n'existe pas."""
        new_dir = tmp_path / "new_output"
        gen = ExcelGenerator(output_path=new_dir)
        gen.ensure_output_dir()
        assert new_dir.exists()

    def test_ensure_output_dir_idempotent(self, generator: ExcelGenerator) -> None:
        """ensure_output_dir peut etre appele plusieurs fois."""
        generator.ensure_output_dir()
        generator.ensure_output_dir()
        assert generator._output_path.exists()


class TestExcelGeneratorErrors:
    """Tests d'erreur du generateur."""

    @pytest.mark.asyncio
    async def test_output_too_large(self, tmp_path: Path) -> None:
        """Fichier trop volumineux leve ExcelGenerationError."""
        gen = ExcelGenerator(output_path=tmp_path)
        # Creer un fichier avec beaucoup de donnees
        # On triche en forçant une taille max tres petite
        gen._max_output_size_mb = 0  # type: ignore[attr-defined]
        params = CreateExcelParams(
            filename="big.xlsx",
            sheets=[
                SheetDef(
                    name="Data",
                    rows=[["x" * 100] * 10 for _ in range(100)],
                )
            ],
        )
        with pytest.raises(ExcelGenerationError):
            await gen.generate(params)

    @pytest.mark.asyncio
    async def test_headers_bold_by_default(self, generator: ExcelGenerator) -> None:
        """Les headers sont automatiquement en gras."""
        params = CreateExcelParams(
            filename="headers.xlsx",
            sheets=[
                SheetDef(
                    name="Test",
                    headers=["H1", "H2"],
                    rows=[["a", "b"]],
                )
            ],
        )
        result = await generator.generate(params)

        wb = load_workbook(result.file_path)
        ws = wb.active
        assert ws is not None
        assert ws["A1"].font.bold is True
        assert ws["B1"].font.bold is True
        wb.close()

    @pytest.mark.asyncio
    async def test_tab_color_applied(self, generator: ExcelGenerator) -> None:
        """Couleur d'onglet appliquee."""
        params = CreateExcelParams(
            filename="tab.xlsx",
            sheets=[SheetDef(name="Colored", tab_color="FF5733")],
        )
        result = await generator.generate(params)

        wb = load_workbook(result.file_path)
        ws = wb["Colored"]
        assert ws.sheet_properties.tabColor is not None
        wb.close()


class TestExcelGeneratorStyles:
    """Tests application des styles."""

    @pytest.mark.asyncio
    async def test_bold_and_italic(self, generator: ExcelGenerator) -> None:
        """Bold et italic appliques."""
        params = CreateExcelParams(
            filename="style_bold.xlsx",
            sheets=[
                SheetDef(
                    name="S1",
                    rows=[["data"]],
                    styles=[CellStyleDef(range="A1", bold=True, italic=True)],
                )
            ],
        )
        result = await generator.generate(params)
        wb = load_workbook(result.file_path)
        ws = wb.active
        assert ws is not None
        assert ws["A1"].font.bold is True
        assert ws["A1"].font.italic is True
        wb.close()

    @pytest.mark.asyncio
    async def test_font_and_bg_color(self, generator: ExcelGenerator) -> None:
        """Couleurs de police et fond appliquees."""
        params = CreateExcelParams(
            filename="style_color.xlsx",
            sheets=[
                SheetDef(
                    name="S1",
                    rows=[["data"]],
                    styles=[
                        CellStyleDef(
                            range="A1",
                            font_color="FF0000",
                            bg_color="4472C4",
                        )
                    ],
                )
            ],
        )
        result = await generator.generate(params)
        wb = load_workbook(result.file_path)
        ws = wb.active
        assert ws is not None
        assert ws["A1"].fill.start_color.rgb.endswith("4472C4")
        wb.close()

    @pytest.mark.asyncio
    async def test_number_format(self, generator: ExcelGenerator) -> None:
        """Format numerique applique."""
        params = CreateExcelParams(
            filename="style_numfmt.xlsx",
            sheets=[
                SheetDef(
                    name="S1",
                    rows=[[1234.56]],
                    styles=[CellStyleDef(range="A1", number_format="#,##0.00")],
                )
            ],
        )
        result = await generator.generate(params)
        wb = load_workbook(result.file_path)
        ws = wb.active
        assert ws is not None
        assert ws["A1"].number_format == "#,##0.00"
        wb.close()

    @pytest.mark.asyncio
    async def test_alignment_center(self, generator: ExcelGenerator) -> None:
        """Alignement centre applique."""
        params = CreateExcelParams(
            filename="style_align.xlsx",
            sheets=[
                SheetDef(
                    name="S1",
                    rows=[["data"]],
                    styles=[CellStyleDef(range="A1", alignment="center")],
                )
            ],
        )
        result = await generator.generate(params)
        wb = load_workbook(result.file_path)
        ws = wb.active
        assert ws is not None
        assert ws["A1"].alignment.horizontal == "center"
        wb.close()

    @pytest.mark.asyncio
    async def test_wrap_text(self, generator: ExcelGenerator) -> None:
        """Retour a la ligne applique."""
        params = CreateExcelParams(
            filename="style_wrap.xlsx",
            sheets=[
                SheetDef(
                    name="S1",
                    rows=[["long text"]],
                    styles=[CellStyleDef(range="A1", wrap_text=True)],
                )
            ],
        )
        result = await generator.generate(params)
        wb = load_workbook(result.file_path)
        ws = wb.active
        assert ws is not None
        assert ws["A1"].alignment.wrap_text is True
        wb.close()

    @pytest.mark.asyncio
    async def test_border(self, generator: ExcelGenerator) -> None:
        """Bordures appliquees."""
        params = CreateExcelParams(
            filename="style_border.xlsx",
            sheets=[
                SheetDef(
                    name="S1",
                    rows=[["data"]],
                    styles=[CellStyleDef(range="A1", border=True)],
                )
            ],
        )
        result = await generator.generate(params)
        wb = load_workbook(result.file_path)
        ws = wb.active
        assert ws is not None
        assert ws["A1"].border.left.style == "thin"
        assert ws["A1"].border.right.style == "thin"
        wb.close()

    @pytest.mark.asyncio
    async def test_style_on_range(self, generator: ExcelGenerator) -> None:
        """Style applique sur une plage de cellules."""
        params = CreateExcelParams(
            filename="style_range.xlsx",
            sheets=[
                SheetDef(
                    name="S1",
                    rows=[["a", "b"], ["c", "d"]],
                    styles=[CellStyleDef(range="A1:B2", bold=True)],
                )
            ],
        )
        result = await generator.generate(params)
        wb = load_workbook(result.file_path)
        ws = wb.active
        assert ws is not None
        for cell_ref in ("A1", "B1", "A2", "B2"):
            assert ws[cell_ref].font.bold is True
        wb.close()


class TestExcelGeneratorCharts:
    """Tests graphiques."""

    @pytest.mark.asyncio
    async def test_bar_chart(self, generator: ExcelGenerator) -> None:
        """Graphique a barres cree."""
        params = CreateExcelParams(
            filename="chart_bar.xlsx",
            sheets=[
                SheetDef(
                    name="Data",
                    headers=["Cat", "Val"],
                    rows=[["A", 10], ["B", 20]],
                    charts=[
                        ChartDef(
                            type="bar",
                            data_range="B1:B3",
                            categories_range="A2:A3",
                            title="Test Bar",
                            position="D1",
                        )
                    ],
                )
            ],
        )
        result = await generator.generate(params)
        assert result.total_charts == 1

        wb = load_workbook(result.file_path)
        ws = wb["Data"]
        assert len(ws._charts) == 1
        wb.close()

    @pytest.mark.asyncio
    async def test_line_chart(self, generator: ExcelGenerator) -> None:
        """Graphique en courbes cree."""
        params = CreateExcelParams(
            filename="chart_line.xlsx",
            sheets=[
                SheetDef(
                    name="Data",
                    headers=["Mois", "Ventes"],
                    rows=[["Jan", 100], ["Fev", 150]],
                    charts=[ChartDef(type="line", data_range="B1:B3", position="D1")],
                )
            ],
        )
        result = await generator.generate(params)
        assert result.total_charts == 1

    @pytest.mark.asyncio
    async def test_pie_chart(self, generator: ExcelGenerator) -> None:
        """Graphique camembert cree."""
        params = CreateExcelParams(
            filename="chart_pie.xlsx",
            sheets=[
                SheetDef(
                    name="Data",
                    headers=["Cat", "Part"],
                    rows=[["A", 40], ["B", 60]],
                    charts=[ChartDef(type="pie", data_range="B1:B3", position="D1")],
                )
            ],
        )
        result = await generator.generate(params)
        assert result.total_charts == 1

    @pytest.mark.asyncio
    async def test_column_chart(self, generator: ExcelGenerator) -> None:
        """Graphique en colonnes (BarChart type=col)."""
        params = CreateExcelParams(
            filename="chart_col.xlsx",
            sheets=[
                SheetDef(
                    name="Data",
                    headers=["Cat", "Val"],
                    rows=[["A", 10], ["B", 20]],
                    charts=[ChartDef(type="column", data_range="B1:B3", position="D1")],
                )
            ],
        )
        result = await generator.generate(params)
        assert result.total_charts == 1

    @pytest.mark.asyncio
    async def test_chart_with_axis_titles(self, generator: ExcelGenerator) -> None:
        """Graphique avec titres d'axes."""
        params = CreateExcelParams(
            filename="chart_axes.xlsx",
            sheets=[
                SheetDef(
                    name="Data",
                    headers=["X", "Y"],
                    rows=[[1, 10], [2, 20]],
                    charts=[
                        ChartDef(
                            type="bar",
                            data_range="B1:B3",
                            position="D1",
                            x_axis_title="Axe X",
                            y_axis_title="Axe Y",
                        )
                    ],
                )
            ],
        )
        result = await generator.generate(params)
        assert result.total_charts == 1

    @pytest.mark.asyncio
    async def test_multiple_charts_counted(self, generator: ExcelGenerator) -> None:
        """Plusieurs graphiques comptabilises."""
        params = CreateExcelParams(
            filename="chart_multi.xlsx",
            sheets=[
                SheetDef(
                    name="Data",
                    headers=["Cat", "V1", "V2"],
                    rows=[["A", 10, 20], ["B", 30, 40]],
                    charts=[
                        ChartDef(type="bar", data_range="B1:B3", position="E1"),
                        ChartDef(type="line", data_range="C1:C3", position="E15"),
                    ],
                )
            ],
        )
        result = await generator.generate(params)
        assert result.total_charts == 2

    @pytest.mark.asyncio
    async def test_chart_failure_graceful_degradation(self, generator: ExcelGenerator) -> None:
        """Un graphique invalide est ignore sans bloquer la generation."""
        params = CreateExcelParams(
            filename="chart_fail.xlsx",
            sheets=[
                SheetDef(
                    name="Data",
                    headers=["Cat", "Val"],
                    rows=[["A", 10], ["B", 20]],
                    charts=[
                        ChartDef(
                            type="bar",
                            data_range="Z99:Z100",  # Range invalide
                            position="D1",
                            title="Broken Chart",
                        ),
                        ChartDef(
                            type="bar",
                            data_range="B1:B3",
                            position="D15",
                            title="Valid Chart",
                        ),
                    ],
                )
            ],
        )
        result = await generator.generate(params)
        # Le fichier est genere malgre l'echec du premier graphique
        assert result.sheets_created == 1
        assert result.total_rows == 2
        # Seul le graphique valide est compte
        assert result.total_charts >= 1

    @pytest.mark.asyncio
    async def test_add_chart_raises_excel_chart_error(self, generator: ExcelGenerator) -> None:
        """add_chart leve ExcelChartError (pas Exception generique)."""
        from unittest.mock import MagicMock

        ws = MagicMock()
        ws.title = "Test"
        chart_def = ChartDef(
            type="bar",
            data_range="INVALID",
            position="A1",
            title="Bad Chart",
        )
        with pytest.raises(ExcelChartError, match="Bad Chart"):
            generator.add_chart(ws, chart_def)


class TestExcelGeneratorValidations:
    """Tests validation de donnees."""

    @pytest.mark.asyncio
    async def test_list_validation(self, generator: ExcelGenerator) -> None:
        """Validation liste deroulante."""
        params = CreateExcelParams(
            filename="dv_list.xlsx",
            sheets=[
                SheetDef(
                    name="S1",
                    rows=[[""]],
                    data_validations=[
                        DataValidationDef(
                            range="A1:A100",
                            type="list",
                            values=["Oui", "Non"],
                        )
                    ],
                )
            ],
        )
        result = await generator.generate(params)

        wb = load_workbook(result.file_path)
        ws = wb.active
        assert ws is not None
        assert len(ws.data_validations.dataValidation) == 1
        wb.close()

    @pytest.mark.asyncio
    async def test_whole_number_validation(self, generator: ExcelGenerator) -> None:
        """Validation nombre entier."""
        params = CreateExcelParams(
            filename="dv_whole.xlsx",
            sheets=[
                SheetDef(
                    name="S1",
                    rows=[[0]],
                    data_validations=[
                        DataValidationDef(
                            range="A1:A10",
                            type="whole",
                            operator="greaterThan",
                            formula1="0",
                        )
                    ],
                )
            ],
        )
        result = await generator.generate(params)
        wb = load_workbook(result.file_path)
        ws = wb.active
        assert ws is not None
        assert len(ws.data_validations.dataValidation) == 1
        wb.close()


class TestExcelGeneratorMergedCells:
    """Tests cellules fusionnees."""

    @pytest.mark.asyncio
    async def test_merge_cells(self, generator: ExcelGenerator) -> None:
        """Fusion de cellules."""
        params = CreateExcelParams(
            filename="merge.xlsx",
            sheets=[
                SheetDef(
                    name="S1",
                    rows=[["", "", ""], ["a", "b", "c"]],
                    merged_cells=[MergedCellDef(range="A1:C1", value="Titre")],
                )
            ],
        )
        result = await generator.generate(params)

        wb = load_workbook(result.file_path)
        ws = wb.active
        assert ws is not None
        assert ws["A1"].value == "Titre"
        assert len(ws.merged_cells.ranges) == 1
        wb.close()

    @pytest.mark.asyncio
    async def test_merge_without_value(self, generator: ExcelGenerator) -> None:
        """Fusion sans valeur."""
        params = CreateExcelParams(
            filename="merge_noval.xlsx",
            sheets=[
                SheetDef(
                    name="S1",
                    rows=[["", ""], ["a", "b"]],
                    merged_cells=[MergedCellDef(range="A1:B1")],
                )
            ],
        )
        result = await generator.generate(params)

        wb = load_workbook(result.file_path)
        ws = wb.active
        assert ws is not None
        assert len(ws.merged_cells.ranges) == 1
        wb.close()


class TestExcelGeneratorFormulaInjection:
    """Tests securite: detection de formules dangereuses dans les cellules."""

    @pytest.mark.asyncio
    async def test_safe_formulas_accepted(self, generator: ExcelGenerator) -> None:
        """Les formules Excel standard ne sont pas bloquees."""
        params = CreateExcelParams(
            filename="safe_formulas.xlsx",
            sheets=[
                SheetDef(
                    name="S1",
                    headers=["Val", "Total"],
                    rows=[
                        [10, "=SUM(A1:A5)"],
                        [20, "=AVERAGE(A1:A5)"],
                        [30, "=IF(A1>0,A1,0)"],
                        [40, "=VLOOKUP(A1,A1:B5,2,FALSE)"],
                        [50, "=COUNT(A1:A5)"],
                    ],
                )
            ],
        )
        result = await generator.generate(params)
        assert result.total_rows == 5

    @pytest.mark.asyncio
    async def test_non_string_values_accepted(self, generator: ExcelGenerator) -> None:
        """Les valeurs non-string (int, float, bool, None) passent sans verification."""
        params = CreateExcelParams(
            filename="non_string.xlsx",
            sheets=[
                SheetDef(
                    name="S1",
                    rows=[[42, 3.14, True, None]],
                )
            ],
        )
        result = await generator.generate(params)
        assert result.total_rows == 1

    @pytest.mark.asyncio
    async def test_dde_injection_blocked(self, generator: ExcelGenerator) -> None:
        """Les payloads DDE sont bloques."""
        params = CreateExcelParams(
            filename="dde.xlsx",
            sheets=[
                SheetDef(
                    name="S1",
                    rows=[["+cmd|'/C calc'"]],
                )
            ],
        )
        with pytest.raises(ExcelFormulaInjectionError, match="cmd"):
            await generator.generate(params)

    @pytest.mark.asyncio
    async def test_cmd_formula_blocked(self, generator: ExcelGenerator) -> None:
        """=CMD() est bloque."""
        params = CreateExcelParams(
            filename="cmd.xlsx",
            sheets=[
                SheetDef(
                    name="S1",
                    rows=[['=CMD("calc")']],
                )
            ],
        )
        with pytest.raises(ExcelFormulaInjectionError):
            await generator.generate(params)

    @pytest.mark.asyncio
    async def test_system_formula_blocked(self, generator: ExcelGenerator) -> None:
        """=SYSTEM() est bloque."""
        params = CreateExcelParams(
            filename="system.xlsx",
            sheets=[
                SheetDef(
                    name="S1",
                    rows=[['=SYSTEM("whoami")']],
                )
            ],
        )
        with pytest.raises(ExcelFormulaInjectionError):
            await generator.generate(params)

    @pytest.mark.asyncio
    async def test_exec_formula_blocked(self, generator: ExcelGenerator) -> None:
        """=EXEC() est bloque."""
        params = CreateExcelParams(
            filename="exec.xlsx",
            sheets=[
                SheetDef(
                    name="S1",
                    rows=[['=EXEC("rm -rf /")']],
                )
            ],
        )
        with pytest.raises(ExcelFormulaInjectionError):
            await generator.generate(params)

    @pytest.mark.asyncio
    async def test_minus_cmd_pipe_blocked(self, generator: ExcelGenerator) -> None:
        """-cmd|'/C calc' est bloque."""
        params = CreateExcelParams(
            filename="minus_cmd.xlsx",
            sheets=[
                SheetDef(
                    name="S1",
                    rows=[["-cmd|'/C calc'"]],
                )
            ],
        )
        with pytest.raises(ExcelFormulaInjectionError):
            await generator.generate(params)

    @pytest.mark.asyncio
    async def test_at_sum_cmd_blocked(self, generator: ExcelGenerator) -> None:
        """@SUM(cmd) pattern bloque."""
        params = CreateExcelParams(
            filename="at_cmd.xlsx",
            sheets=[
                SheetDef(
                    name="S1",
                    rows=[["@SUM(1+1)*cmd|'/C calc'"]],
                )
            ],
        )
        with pytest.raises(ExcelFormulaInjectionError):
            await generator.generate(params)

    @pytest.mark.asyncio
    async def test_call_formula_blocked(self, generator: ExcelGenerator) -> None:
        """=CALL() est bloque."""
        params = CreateExcelParams(
            filename="call.xlsx",
            sheets=[
                SheetDef(
                    name="S1",
                    rows=[['=CALL("kernel32","WinExec","JCJ","calc.exe",5)']],
                )
            ],
        )
        with pytest.raises(ExcelFormulaInjectionError):
            await generator.generate(params)

    @pytest.mark.asyncio
    async def test_header_injection_blocked(self, generator: ExcelGenerator) -> None:
        """Les headers sont aussi verifies."""
        params = CreateExcelParams(
            filename="header_inj.xlsx",
            sheets=[
                SheetDef(
                    name="S1",
                    headers=["Name", '=CMD("calc")'],
                    rows=[["a", "b"]],
                )
            ],
        )
        with pytest.raises(ExcelFormulaInjectionError):
            await generator.generate(params)

    @pytest.mark.asyncio
    async def test_case_insensitive_detection(self, generator: ExcelGenerator) -> None:
        """La detection est insensible a la casse."""
        params = CreateExcelParams(
            filename="case.xlsx",
            sheets=[
                SheetDef(
                    name="S1",
                    rows=[['=Cmd("calc")']],
                )
            ],
        )
        with pytest.raises(ExcelFormulaInjectionError):
            await generator.generate(params)
