# SPDX-License-Identifier: AGPL-3.0-or-later
# Copyright (C) 2026 Assistance Micro Design
"""Tests pour le service ExcelEditor."""

from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import load_workbook

from src.core.exceptions import (
    ExcelFileNotFoundError,
    ExcelFormulaInjectionError,
    ExcelGenerationError,
    ExcelSheetNotFoundError,
)
from src.models.api import CreateExcelParams, EditExcelParams
from src.models.excel_edit import (
    AddChartOp,
    AddDataValidationOp,
    AddSheetOp,
    ApplyStylesOp,
    DeleteRowsOp,
    DeleteSheetOp,
    InsertRowsOp,
    MergeCellsOp,
    RemoveChartsOp,
    RenameSheetOp,
    SetSheetPropertiesOp,
    UnmergeCellsOp,
    UpdateCellsOp,
)
from src.models.excel_generation import (
    CellStyleDef,
    ChartDef,
    DataValidationDef,
    MergedCellDef,
    SheetDef,
)
from src.services.excel.editor import ExcelEditor
from src.services.excel.generator import ExcelGenerator


@pytest.fixture
def generator(tmp_path: Path) -> ExcelGenerator:
    """Generateur pour creer les fichiers de test."""
    return ExcelGenerator(output_path=tmp_path)


@pytest.fixture
def editor(tmp_path: Path) -> ExcelEditor:
    """Editeur avec output dans tmp_path."""
    return ExcelEditor(output_path=tmp_path)


@pytest.fixture
async def sample_file(generator: ExcelGenerator) -> str:
    """Cree un fichier Excel de test et retourne le filename."""
    params = CreateExcelParams(
        filename="sample.xlsx",
        sheets=[
            SheetDef(
                name="Ventes",
                headers=["Produit", "Prix", "Quantite"],
                rows=[["Widget", 10, 100], ["Gadget", 20, 50]],
            ),
            SheetDef(name="Resume", rows=[["Total", 1500]]),
        ],
    )
    await generator.generate(params)
    return "sample.xlsx"


# === Sprint 2: Operations basiques ===


class TestUpdateCells:
    """Tests operation update_cells."""

    @pytest.mark.asyncio
    async def test_update_single_cell(self, editor: ExcelEditor, sample_file: str) -> None:
        """Modifier une seule cellule."""
        params = EditExcelParams(
            filename=sample_file,
            operations=[UpdateCellsOp(sheet="Ventes", cells={"B2": 15})],
        )
        result = await editor.edit(params)

        assert result.operations_applied == 1
        wb = load_workbook(result.file_path)
        assert wb["Ventes"]["B2"].value == 15
        wb.close()

    @pytest.mark.asyncio
    async def test_update_multiple_cells(self, editor: ExcelEditor, sample_file: str) -> None:
        """Modifier plusieurs cellules."""
        params = EditExcelParams(
            filename=sample_file,
            operations=[
                UpdateCellsOp(sheet="Ventes", cells={"A2": "SuperWidget", "B2": 25, "C2": 200})
            ],
        )
        result = await editor.edit(params)

        wb = load_workbook(result.file_path)
        ws = wb["Ventes"]
        assert ws["A2"].value == "SuperWidget"
        assert ws["B2"].value == 25
        assert ws["C2"].value == 200
        wb.close()

    @pytest.mark.asyncio
    async def test_update_with_formula(self, editor: ExcelEditor, sample_file: str) -> None:
        """Ecrire une formule dans une cellule."""
        params = EditExcelParams(
            filename=sample_file,
            operations=[UpdateCellsOp(sheet="Ventes", cells={"D2": "=B2*C2"})],
        )
        result = await editor.edit(params)

        wb = load_workbook(result.file_path)
        assert wb["Ventes"]["D2"].value == "=B2*C2"
        wb.close()


class TestInsertRows:
    """Tests operation insert_rows."""

    @pytest.mark.asyncio
    async def test_append_rows(self, editor: ExcelEditor, sample_file: str) -> None:
        """Ajouter des lignes a la fin."""
        params = EditExcelParams(
            filename=sample_file,
            operations=[InsertRowsOp(sheet="Ventes", rows=[["Nouveau", 30, 75]])],
        )
        result = await editor.edit(params)

        wb = load_workbook(result.file_path)
        ws = wb["Ventes"]
        # Headers=row1, 2 lignes existantes=row2-3, nouvelle=row4
        assert ws["A4"].value == "Nouveau"
        assert ws["B4"].value == 30
        wb.close()

    @pytest.mark.asyncio
    async def test_insert_rows_at_position(self, editor: ExcelEditor, sample_file: str) -> None:
        """Inserer des lignes a une position specifique."""
        params = EditExcelParams(
            filename=sample_file,
            operations=[InsertRowsOp(sheet="Ventes", rows=[["Inserted", 5, 10]], at_row=2)],
        )
        result = await editor.edit(params)

        wb = load_workbook(result.file_path)
        ws = wb["Ventes"]
        # La ligne 2 originale (Widget) est decalee en 3
        assert ws["A2"].value == "Inserted"
        assert ws["A3"].value == "Widget"
        wb.close()


class TestDeleteRows:
    """Tests operation delete_rows."""

    @pytest.mark.asyncio
    async def test_delete_single_row(self, editor: ExcelEditor, sample_file: str) -> None:
        """Supprimer une seule ligne."""
        params = EditExcelParams(
            filename=sample_file,
            operations=[DeleteRowsOp(sheet="Ventes", start_row=2, end_row=2)],
        )
        result = await editor.edit(params)

        wb = load_workbook(result.file_path)
        ws = wb["Ventes"]
        # Ligne 2 (Widget) supprimee, Gadget remonte en row 2
        assert ws["A2"].value == "Gadget"
        wb.close()

    @pytest.mark.asyncio
    async def test_delete_multiple_rows(self, editor: ExcelEditor, sample_file: str) -> None:
        """Supprimer plusieurs lignes."""
        params = EditExcelParams(
            filename=sample_file,
            operations=[DeleteRowsOp(sheet="Ventes", start_row=2, end_row=3)],
        )
        result = await editor.edit(params)

        wb = load_workbook(result.file_path)
        ws = wb["Ventes"]
        # Les 2 lignes de donnees supprimees, seule la row 1 (headers) reste
        assert ws["A1"].value == "Produit"
        assert ws["A2"].value is None
        wb.close()


class TestApplyStyles:
    """Tests operation apply_styles."""

    @pytest.mark.asyncio
    async def test_apply_bold(self, editor: ExcelEditor, sample_file: str) -> None:
        """Appliquer bold sur une cellule."""
        params = EditExcelParams(
            filename=sample_file,
            operations=[
                ApplyStylesOp(sheet="Ventes", styles=[CellStyleDef(range="A2", bold=True)])
            ],
        )
        result = await editor.edit(params)

        wb = load_workbook(result.file_path)
        assert wb["Ventes"]["A2"].font.bold is True
        wb.close()

    @pytest.mark.asyncio
    async def test_apply_bg_color(self, editor: ExcelEditor, sample_file: str) -> None:
        """Appliquer couleur de fond."""
        params = EditExcelParams(
            filename=sample_file,
            operations=[
                ApplyStylesOp(
                    sheet="Ventes",
                    styles=[CellStyleDef(range="A1:C1", bg_color="4472C4")],
                )
            ],
        )
        result = await editor.edit(params)

        wb = load_workbook(result.file_path)
        assert wb["Ventes"]["A1"].fill.start_color.rgb.endswith("4472C4")
        wb.close()

    @pytest.mark.asyncio
    async def test_apply_border(self, editor: ExcelEditor, sample_file: str) -> None:
        """Appliquer bordures."""
        params = EditExcelParams(
            filename=sample_file,
            operations=[
                ApplyStylesOp(
                    sheet="Ventes",
                    styles=[CellStyleDef(range="A1", border=True)],
                )
            ],
        )
        result = await editor.edit(params)

        wb = load_workbook(result.file_path)
        assert wb["Ventes"]["A1"].border.left.style == "thin"
        wb.close()


# === Sprint 3: Sheets + Charts ===


class TestAddSheet:
    """Tests operation add_sheet."""

    @pytest.mark.asyncio
    async def test_add_empty_sheet(self, editor: ExcelEditor, sample_file: str) -> None:
        """Ajouter une feuille vide."""
        params = EditExcelParams(
            filename=sample_file,
            operations=[AddSheetOp(name="Nouvelle")],
        )
        result = await editor.edit(params)

        wb = load_workbook(result.file_path)
        assert "Nouvelle" in wb.sheetnames
        wb.close()

    @pytest.mark.asyncio
    async def test_add_sheet_with_data(self, editor: ExcelEditor, sample_file: str) -> None:
        """Ajouter une feuille avec headers et donnees."""
        params = EditExcelParams(
            filename=sample_file,
            operations=[
                AddSheetOp(
                    name="Stocks",
                    headers=["Item", "Qty"],
                    rows=[["A", 10], ["B", 20]],
                )
            ],
        )
        result = await editor.edit(params)

        wb = load_workbook(result.file_path)
        ws = wb["Stocks"]
        assert ws["A1"].value == "Item"
        assert ws["A2"].value == "A"
        assert ws["B2"].value == 10
        wb.close()


class TestDeleteSheet:
    """Tests operation delete_sheet."""

    @pytest.mark.asyncio
    async def test_delete_sheet(self, editor: ExcelEditor, sample_file: str) -> None:
        """Supprimer une feuille."""
        params = EditExcelParams(
            filename=sample_file,
            operations=[DeleteSheetOp(name="Resume")],
        )
        result = await editor.edit(params)

        wb = load_workbook(result.file_path)
        assert "Resume" not in wb.sheetnames
        assert "Ventes" in wb.sheetnames
        wb.close()


class TestRenameSheet:
    """Tests operation rename_sheet."""

    @pytest.mark.asyncio
    async def test_rename_sheet(self, editor: ExcelEditor, sample_file: str) -> None:
        """Renommer une feuille."""
        params = EditExcelParams(
            filename=sample_file,
            operations=[RenameSheetOp(name="Resume", new_name="Synthese")],
        )
        result = await editor.edit(params)

        wb = load_workbook(result.file_path)
        assert "Synthese" in wb.sheetnames
        assert "Resume" not in wb.sheetnames
        wb.close()


class TestAddChart:
    """Tests operation add_chart."""

    @pytest.mark.asyncio
    async def test_add_bar_chart(self, editor: ExcelEditor, sample_file: str) -> None:
        """Ajouter un graphique a barres."""
        params = EditExcelParams(
            filename=sample_file,
            operations=[
                AddChartOp(
                    sheet="Ventes",
                    chart=ChartDef(
                        type="bar",
                        data_range="B1:B3",
                        categories_range="A2:A3",
                        title="Ventes par produit",
                        position="E1",
                    ),
                )
            ],
        )
        result = await editor.edit(params)

        wb = load_workbook(result.file_path)
        assert len(wb["Ventes"]._charts) == 1
        wb.close()


class TestRemoveCharts:
    """Tests operation remove_charts."""

    @pytest.mark.asyncio
    async def test_remove_charts(self, editor: ExcelEditor, sample_file: str) -> None:
        """Supprimer tous les graphiques d'une feuille."""
        # D'abord ajouter un chart
        await editor.edit(
            EditExcelParams(
                filename=sample_file,
                operations=[
                    AddChartOp(
                        sheet="Ventes",
                        chart=ChartDef(type="bar", data_range="B1:B3", position="E1"),
                    )
                ],
            )
        )

        # Puis supprimer
        result = await editor.edit(
            EditExcelParams(
                filename=sample_file,
                operations=[RemoveChartsOp(sheet="Ventes")],
            )
        )

        wb = load_workbook(result.file_path)
        assert len(wb["Ventes"]._charts) == 0
        wb.close()


# === Sprint 4: Validations, Merges, Properties, Erreurs ===


class TestAddDataValidation:
    """Tests operation add_data_validation."""

    @pytest.mark.asyncio
    async def test_add_list_validation(self, editor: ExcelEditor, sample_file: str) -> None:
        """Ajouter une validation liste."""
        params = EditExcelParams(
            filename=sample_file,
            operations=[
                AddDataValidationOp(
                    sheet="Ventes",
                    validation=DataValidationDef(
                        range="D2:D100",
                        type="list",
                        values=["Oui", "Non", "Peut-etre"],
                    ),
                )
            ],
        )
        result = await editor.edit(params)

        wb = load_workbook(result.file_path)
        assert len(wb["Ventes"].data_validations.dataValidation) >= 1
        wb.close()


class TestMergeCells:
    """Tests operation merge_cells."""

    @pytest.mark.asyncio
    async def test_merge_with_value(self, editor: ExcelEditor, sample_file: str) -> None:
        """Fusionner des cellules avec valeur."""
        params = EditExcelParams(
            filename=sample_file,
            operations=[
                MergeCellsOp(
                    sheet="Resume",
                    merge=MergedCellDef(range="A1:B1", value="Total General"),
                )
            ],
        )
        result = await editor.edit(params)

        wb = load_workbook(result.file_path)
        ws = wb["Resume"]
        assert ws["A1"].value == "Total General"
        assert len(ws.merged_cells.ranges) == 1
        wb.close()


class TestUnmergeCells:
    """Tests operation unmerge_cells."""

    @pytest.mark.asyncio
    async def test_unmerge(self, editor: ExcelEditor, sample_file: str) -> None:
        """Defusionner des cellules."""
        # D'abord fusionner
        await editor.edit(
            EditExcelParams(
                filename=sample_file,
                operations=[
                    MergeCellsOp(
                        sheet="Resume",
                        merge=MergedCellDef(range="A1:B1"),
                    )
                ],
            )
        )

        # Puis defusionner
        result = await editor.edit(
            EditExcelParams(
                filename=sample_file,
                operations=[UnmergeCellsOp(sheet="Resume", range="A1:B1")],
            )
        )

        wb = load_workbook(result.file_path)
        assert len(wb["Resume"].merged_cells.ranges) == 0
        wb.close()


class TestSetSheetProperties:
    """Tests operation set_sheet_properties."""

    @pytest.mark.asyncio
    async def test_set_column_widths(self, editor: ExcelEditor, sample_file: str) -> None:
        """Modifier largeurs de colonnes."""
        params = EditExcelParams(
            filename=sample_file,
            operations=[SetSheetPropertiesOp(sheet="Ventes", column_widths={"A": 30, "B": 15})],
        )
        result = await editor.edit(params)

        wb = load_workbook(result.file_path)
        ws = wb["Ventes"]
        assert ws.column_dimensions["A"].width == 30
        assert ws.column_dimensions["B"].width == 15
        wb.close()

    @pytest.mark.asyncio
    async def test_set_freeze_panes(self, editor: ExcelEditor, sample_file: str) -> None:
        """Modifier gel de volets."""
        params = EditExcelParams(
            filename=sample_file,
            operations=[SetSheetPropertiesOp(sheet="Ventes", freeze_panes="A2")],
        )
        result = await editor.edit(params)

        wb = load_workbook(result.file_path)
        assert wb["Ventes"].freeze_panes == "A2"
        wb.close()

    @pytest.mark.asyncio
    async def test_set_tab_color(self, editor: ExcelEditor, sample_file: str) -> None:
        """Modifier couleur d'onglet."""
        params = EditExcelParams(
            filename=sample_file,
            operations=[SetSheetPropertiesOp(sheet="Ventes", tab_color="FF5733")],
        )
        result = await editor.edit(params)

        wb = load_workbook(result.file_path)
        assert wb["Ventes"].sheet_properties.tabColor is not None
        wb.close()

    @pytest.mark.asyncio
    async def test_set_auto_filter(self, editor: ExcelEditor, sample_file: str) -> None:
        """Activer auto-filtre."""
        params = EditExcelParams(
            filename=sample_file,
            operations=[SetSheetPropertiesOp(sheet="Ventes", auto_filter=True)],
        )
        result = await editor.edit(params)

        wb = load_workbook(result.file_path)
        assert wb["Ventes"].auto_filter.ref is not None
        wb.close()


# === Erreurs ===


class TestEditorErrors:
    """Tests des cas d'erreur."""

    @pytest.mark.asyncio
    async def test_file_not_found(self, editor: ExcelEditor) -> None:
        """Fichier inexistant leve ExcelFileNotFoundError."""
        params = EditExcelParams(
            filename="inexistant.xlsx",
            operations=[UpdateCellsOp(sheet="S1", cells={"A1": 1})],
        )
        with pytest.raises(ExcelFileNotFoundError):
            await editor.edit(params)

    @pytest.mark.asyncio
    async def test_sheet_not_found(self, editor: ExcelEditor, sample_file: str) -> None:
        """Feuille inexistante leve ExcelSheetNotFoundError."""
        params = EditExcelParams(
            filename=sample_file,
            operations=[UpdateCellsOp(sheet="Inexistante", cells={"A1": 1})],
        )
        with pytest.raises(ExcelSheetNotFoundError):
            await editor.edit(params)

    @pytest.mark.asyncio
    async def test_output_too_large(self, tmp_path: Path, sample_file: str) -> None:
        """Fichier trop gros leve ExcelGenerationError."""
        # Recreer sample_file car le fixture generator utilise tmp_path
        ed = ExcelEditor(output_path=tmp_path)
        ed._generator._max_output_size_mb = 0  # type: ignore[attr-defined]
        params = EditExcelParams(
            filename=sample_file,
            operations=[
                InsertRowsOp(
                    sheet="Ventes",
                    rows=[["x" * 100] * 3 for _ in range(500)],
                )
            ],
        )
        with pytest.raises(ExcelGenerationError):
            await ed.edit(params)


class TestMultipleOperations:
    """Tests operations multiples en sequence."""

    @pytest.mark.asyncio
    async def test_multi_ops_sequence(self, editor: ExcelEditor, sample_file: str) -> None:
        """Plusieurs operations en une seule requete."""
        params = EditExcelParams(
            filename=sample_file,
            operations=[
                UpdateCellsOp(sheet="Ventes", cells={"B2": 99}),
                InsertRowsOp(sheet="Ventes", rows=[["Truc", 5, 10]]),
                ApplyStylesOp(sheet="Ventes", styles=[CellStyleDef(range="A1:C1", bold=True)]),
                RenameSheetOp(name="Resume", new_name="Bilan"),
            ],
        )
        result = await editor.edit(params)

        assert result.operations_applied == 4
        assert result.operations_skipped == 0

        wb = load_workbook(result.file_path)
        assert wb["Ventes"]["B2"].value == 99
        assert "Bilan" in wb.sheetnames
        wb.close()


class TestEditorFormulaInjection:
    """Tests securite: detection de formules dangereuses via editor."""

    @pytest.mark.asyncio
    async def test_update_cells_blocks_dde(self, editor: ExcelEditor, sample_file: str) -> None:
        """update_cells bloque les payloads DDE."""
        params = EditExcelParams(
            filename=sample_file,
            operations=[
                UpdateCellsOp(sheet="Ventes", cells={"A1": "+cmd|'/C calc'"}),
            ],
        )
        with pytest.raises(ExcelFormulaInjectionError):
            await editor.edit(params)

    @pytest.mark.asyncio
    async def test_update_cells_allows_safe_formulas(
        self, editor: ExcelEditor, sample_file: str
    ) -> None:
        """update_cells autorise les formules Excel standard."""
        params = EditExcelParams(
            filename=sample_file,
            operations=[
                UpdateCellsOp(sheet="Ventes", cells={"D1": "=SUM(B2:B3)"}),
            ],
        )
        result = await editor.edit(params)
        assert result.operations_applied == 1

    @pytest.mark.asyncio
    async def test_insert_rows_blocks_injection(
        self, editor: ExcelEditor, sample_file: str
    ) -> None:
        """insert_rows bloque les valeurs dangereuses."""
        params = EditExcelParams(
            filename=sample_file,
            operations=[
                InsertRowsOp(sheet="Ventes", rows=[['=CMD("calc")', 10, 5]]),
            ],
        )
        with pytest.raises(ExcelFormulaInjectionError):
            await editor.edit(params)

    @pytest.mark.asyncio
    async def test_add_sheet_blocks_injection(self, editor: ExcelEditor, sample_file: str) -> None:
        """add_sheet bloque les headers et valeurs dangereuses."""
        params = EditExcelParams(
            filename=sample_file,
            operations=[
                AddSheetOp(
                    name="Evil",
                    headers=['=SYSTEM("whoami")'],
                    rows=[["safe"]],
                ),
            ],
        )
        with pytest.raises(ExcelFormulaInjectionError):
            await editor.edit(params)
